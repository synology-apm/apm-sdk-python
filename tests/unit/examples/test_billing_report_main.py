"""Unit tests for billing_report.py: main() validation/wiring and output functions."""
from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path
from typing import Any
from unittest.mock import patch

import billing_report
import openpyxl
import pytest
from billing_report import (
    _PRICING_TEMPLATE,
    _DetailRow,
    _DetailsView,
    _GroupCharge,
    _GroupPlanRow,
    _GroupServerRow,
    _GroupTypeRow,
    _PlanCharge,
    _PricingConfig,
    _PricingPlan,
    _print_billing_table,
    _print_csv,
    _print_details_tables,
    _print_dist_table,
    _print_json,
    _print_pricing_plans_table,
    _print_table,
    _ServerCharge,
    _ServerDetailRow,
    _write_xlsx,
    main,
)

from tests.unit.examples._billing_fixtures import (
    make_default_config,
    make_group_charge,
    make_plan_charge,
    make_server_charge,
    make_two_plan_config,
)

# ── helpers ─────────────────────────────────────────────────────────────────────


def _argv(*args: str) -> list[str]:
    return ["billing_report.py", *args]


def _line_with(out: str, needle: str) -> str:
    """The single output line containing *needle* (fails if absent or ambiguous)."""
    lines = [line for line in out.splitlines() if needle in line]
    assert len(lines) == 1, f"expected exactly one line containing {needle!r}, got {lines!r}"
    return lines[0]


def _make_detail_scenario() -> tuple[list[_PlanCharge], list[_GroupCharge], list[_ServerCharge], _DetailsView]:
    """One group (Contoso, Premium), one server, one plan, and a coherent details view."""
    plan_charges = [make_plan_charge(instances=2, storage_gb=1.11, instance_charge=10.0, storage_charge=0.22)]
    group_charges = [
        make_group_charge(group_name="Contoso", instances=2, storage_gb=1.11, instance_charge=20.0, storage_charge=0.33)
    ]
    server_charges = [make_server_charge(instances=2, storage_gb=1.11, instance_charge=10.0, storage_charge=0.22)]
    details = _DetailsView(
        server_rows_by_group={
            "Contoso": [_GroupServerRow(
                group_name="Contoso", server_name="apm-server-01", namespace="ns-001",
                instances=2, storage_gb=1.11,
            )],
        },
        plan_rows_by_group={
            "Contoso": [_GroupPlanRow(
                group_name="Contoso", plan_name="Daily Backup", plan_id="plan-001",
                plan_type="Protection Plan", instances=2, storage_gb=1.11,
            )],
        },
        workload_rows_by_group={
            "Contoso": [_GroupTypeRow(
                group_name="Contoso", pricing_plan="Premium", workload_type="VM",
                instances=2, storage_gb=1.1084, instance_charge=20.0, storage_charge=0.33,
            )],
        },
        server_type_rows=[_ServerDetailRow(
            server_name="apm-server-01", namespace="ns-001", pricing_plan="Standard",
            workload_type="VM", instances=2, storage_gb=1.1084,
            instance_charge=10.0, storage_charge=0.22,
        )],
        plan_type_rows=[_DetailRow(
            plan_type="Protection Plan", plan_name="Daily Backup", plan_id="plan-001",
            pricing_plan="Standard", workload_type="VM", instances=2, storage_gb=1.1084,
            instance_charge=10.0, storage_charge=0.22,
        )],
    )
    return plan_charges, group_charges, server_charges, details


# ── --dump-config-template ────────────────────────────────────────────────────


def test_dump_config_template_exits_0(capsys: pytest.CaptureFixture[str]) -> None:
    with patch.object(sys, "argv", _argv("--dump-config-template")):
        with pytest.raises(SystemExit) as exc:
            main()
    assert exc.value.code == 0
    assert capsys.readouterr().out == _PRICING_TEMPLATE


# ── main() error exits ────────────────────────────────────────────────────────

_VALID_CFG = (
    "pricing_plans:\n"
    "  - name: Standard\n"
    "    charge_per_instance: 5.0\n"
    "    charge_per_gb: 0.2\n"
)


@pytest.mark.parametrize("argv,cfg_yaml,exit_code,stderr_match", [
    (["-o", "xlsx"], None, 2, "-o xlsx requires --output-file"),
    (["--config", "{cfg}", "--charge-per-instance", "3.0"], _VALID_CFG, 2, "--config cannot be combined with"),
    (["--show-not-configured"], None, 2, "--show-not-configured requires --config"),
    (["--config", "{cfg}", "--only", "groups"], _VALID_CFG, 2, "requires a --config file that defines groups"),
    (["--config", "{cfg}"], "pricing_plans: []\n", 1, "Error loading pricing file:"),
    (["--config", "/nonexistent/pricing.yaml"], None, 1, "Error loading pricing file:"),
    (["--config", "{cfg}"], "pricing_plans:\n  - charge_per_instance: 5.0\n", 1, "'name'"),
], ids=[
    "xlsx-without-output-file",
    "config-with-flat-rate",
    "show-not-configured-without-config",
    "only-groups-without-groups",
    "invalid-config-file",
    "missing-config-file",
    "pricing-plan-missing-name",
])
def test_main_error_exits(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    argv: list[str],
    cfg_yaml: str | None,
    exit_code: int,
    stderr_match: str,
) -> None:
    if cfg_yaml is not None:
        cfg = tmp_path / "pricing.yaml"
        cfg.write_text(cfg_yaml)
        argv = [a.replace("{cfg}", str(cfg)) for a in argv]
    with patch.object(sys, "argv", _argv(*argv)):
        with pytest.raises(SystemExit) as exc:
            main()
    assert exc.value.code == exit_code
    assert stderr_match in capsys.readouterr().err


# ── main() argument wiring ────────────────────────────────────────────────────


def _capture_run_kwargs(monkeypatch: pytest.MonkeyPatch, argv: list[str]) -> dict[str, Any]:
    """Run main() with run/run_main stubbed out; return the kwargs passed to run()."""
    captured: dict[str, Any] = {}

    def _fake_run(**kwargs: Any) -> None:
        captured.update(kwargs)

    monkeypatch.setattr(billing_report, "run", _fake_run)
    monkeypatch.setattr(billing_report, "run_main", lambda coro: None)
    with patch.object(sys, "argv", _argv(*argv)):
        main()
    return captured


def test_main_flat_rate_builds_unnamed_pricing_plan(monkeypatch: pytest.MonkeyPatch) -> None:
    captured = _capture_run_kwargs(
        monkeypatch, ["--charge-per-instance", "5", "--charge-per-gb", "0.1"]
    )
    pricing = captured["pricing"]
    assert pricing.pricing_plans == [_PricingPlan("", 5.0, 0.1)]
    assert pricing.groups == []
    assert captured["output_format"] == "table"
    assert captured["show_details"] is False
    assert captured["concurrency"] == 5
    assert captured["output_file"] is None
    assert captured["only"] == ""
    assert captured["configured_only"] is False


@pytest.mark.parametrize("extra_argv,expected_configured_only", [
    ([], True),
    (["--show-not-configured"], False),
], ids=["default", "show-not-configured"])
def test_main_config_sets_configured_only(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    extra_argv: list[str],
    expected_configured_only: bool,
) -> None:
    cfg = tmp_path / "pricing.yaml"
    cfg.write_text(_VALID_CFG)
    captured = _capture_run_kwargs(monkeypatch, ["--config", str(cfg), *extra_argv])
    assert captured["pricing"].pricing_plans == [_PricingPlan("Standard", 5.0, 0.2)]
    assert captured["configured_only"] is expected_configured_only


# ── _print_json — direct calls ────────────────────────────────────────────────


def _run_print_json(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    capsys: pytest.CaptureFixture[str],
    only: str = "",
    details: _DetailsView | None = None,
    pricing: _PricingConfig | None = None,
) -> Any:
    if pricing is None:
        pricing = make_default_config()
    _print_json(plan_charges, group_charges, server_charges, pricing, details, only)
    return json.loads(capsys.readouterr().out)


def test_print_json_plan_fields(capsys: pytest.CaptureFixture[str]) -> None:
    charge = make_plan_charge()
    out = _run_print_json([charge], [], [], capsys)

    assert len(out["plans"]) == 1
    plan = out["plans"][0]
    assert plan["plan_name"] == "Daily Backup"
    assert plan["plan_id"] == "plan-001"
    assert plan["instances"] == 3
    assert plan["storage_gb"] == 2.0
    assert plan["total_charge"] == 15.40
    assert plan["pricing_plan"]["name"] == "Standard"
    assert plan["pricing_plan"]["charge_per_instance"] == 5.0


def test_print_json_totals(capsys: pytest.CaptureFixture[str]) -> None:
    c1 = make_plan_charge(
        plan_name="P1", plan_id="plan-001",
        instances=2, storage_gb=1.0, instance_charge=10.0, storage_charge=0.20,
    )
    c2 = make_plan_charge(
        plan_name="P2", plan_id="plan-002",
        instances=3, storage_gb=2.0, instance_charge=15.0, storage_charge=0.40,
    )
    out = _run_print_json([c1, c2], [], [], capsys)

    totals = out["totals"]["plans"]
    assert totals["instances"] == 5
    assert totals["storage_gb"] == 3.0
    assert totals["total_charge"] == 25.60


def test_print_json_only_plans_omits_groups_and_servers(capsys: pytest.CaptureFixture[str]) -> None:
    charge = make_plan_charge()
    out = _run_print_json([charge], [], [], capsys, only="plans")
    assert out["groups"] == []
    assert out["backup_servers"] == []
    assert len(out["plans"]) == 1
    assert "plans" in out["totals"]
    assert "groups" not in out["totals"]
    assert "backup_servers" not in out["totals"]


def test_print_json_empty_plan_list(capsys: pytest.CaptureFixture[str]) -> None:
    out = _run_print_json([], [], [], capsys)
    assert out["plans"] == []
    assert out["groups"] == []
    assert out["backup_servers"] == []
    totals = out["totals"]
    assert totals["plans"]["instances"] == 0
    assert totals["plans"]["total_charge"] == 0.0


def test_print_json_pricing_plans_list(capsys: pytest.CaptureFixture[str]) -> None:
    out = _run_print_json([], [], [], capsys)
    assert len(out["pricing_plans"]) == 1
    assert out["pricing_plans"][0]["name"] == "Standard"
    assert out["pricing_plans"][0]["charge_per_instance"] == 5.0
    assert out["pricing_plans"][0]["charge_per_gb"] == 0.20


def test_print_json_group_and_server_entries(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, _ = _make_detail_scenario()
    out = _run_print_json(plan_charges, group_charges, server_charges, capsys)

    group = out["groups"][0]
    assert group["group_name"] == "Contoso"
    assert group["plan_type"] == "Protection Plan"
    assert group["pricing_plan"]["name"] == "Premium"
    assert group["instances"] == 2
    assert group["total_charge"] == 20.33
    server = out["backup_servers"][0]
    assert server["server_name"] == "apm-server-01"
    assert server["namespace"] == "ns-001"
    assert server["total_charge"] == 10.22
    assert out["totals"]["groups"]["total_charge"] == 20.33
    assert out["totals"]["backup_servers"]["total_charge"] == 10.22


def test_print_json_details_nested_breakdowns(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, details = _make_detail_scenario()
    out = _run_print_json(plan_charges, group_charges, server_charges, capsys, details=details)

    group = out["groups"][0]
    assert group["backup_servers"] == [
        {"server_name": "apm-server-01", "namespace": "ns-001", "instances": 2, "storage_gb": 1.11}
    ]
    assert group["plans"][0]["plan_id"] == "plan-001"
    assert group["plans"][0]["plan_type"] == "Protection Plan"
    assert group["workloads"][0]["workload_type"] == "VM"
    assert group["workloads"][0]["total_charge"] == 20.33
    server_by_type = out["backup_servers"][0]["by_type"][0]
    assert server_by_type["workload_type"] == "VM"
    assert server_by_type["storage_gb"] == 1.1084
    assert server_by_type["total_charge"] == 10.22
    plan_by_type = out["plans"][0]["by_type"][0]
    assert plan_by_type["workload_type"] == "VM"
    assert plan_by_type["instance_charge"] == 10.0


# ── _print_csv — direct calls ──────────────────────────────────────────────────


def _run_print_csv(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    capsys: pytest.CaptureFixture[str],
    only: str = "",
    details: _DetailsView | None = None,
    pricing: _PricingConfig | None = None,
) -> list[list[str]]:
    if pricing is None:
        pricing = make_default_config()
    _print_csv(plan_charges, group_charges, server_charges, pricing, details, only)
    output = capsys.readouterr().out
    return list(csv.reader(io.StringIO(output)))


def test_print_csv_plan_section_header_and_data(capsys: pytest.CaptureFixture[str]) -> None:
    charge = make_plan_charge()
    rows = _run_print_csv([charge], [], [], capsys)

    # First row: section title
    assert rows[0] == ["Plans"]
    # Second row: column headers
    header = rows[1]
    assert "plan" in header
    assert "instances" in header
    assert "total_charge" in header
    # Third row: data
    data = rows[2]
    assert data[header.index("plan")] == "Daily Backup"
    assert data[header.index("instances")] == "3"
    assert data[header.index("total_charge")] == "15.4"
    # Fourth row: Total
    total_row = rows[3]
    assert total_row[0] == "Total"
    assert total_row[header.index("total_charge")] == "15.4"


def test_print_csv_group_and_server_sections(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, _ = _make_detail_scenario()
    rows = _run_print_csv(plan_charges, group_charges, server_charges, capsys)

    assert [r[0] for r in rows if len(r) == 1] == ["Groups", "Backup Servers", "Plans"]
    group_header = rows[1]
    group_data = rows[2]
    assert group_data[group_header.index("group")] == "Contoso"
    assert group_data[group_header.index("kind")] == "Protection Plan"
    assert group_data[group_header.index("pricing_plan")] == "Premium"
    assert group_data[group_header.index("instances")] == "2"
    assert group_data[group_header.index("total_charge")] == "20.33"
    assert rows[3][0] == "Total"
    server_start = rows.index(["Backup Servers"])
    server_header = rows[server_start + 1]
    server_data = rows[server_start + 2]
    assert server_data[server_header.index("server_name")] == "apm-server-01"
    assert server_data[server_header.index("namespace")] == "ns-001"
    assert server_data[server_header.index("total_charge")] == "10.22"


def test_print_csv_empty_plan_list_produces_no_plan_section(
    capsys: pytest.CaptureFixture[str],
) -> None:
    rows = _run_print_csv([], [], [], capsys)
    assert rows == []


def test_print_csv_only_plans(capsys: pytest.CaptureFixture[str]) -> None:
    charge = make_plan_charge()
    rows = _run_print_csv([charge], [], [], capsys, only="plans")
    section_titles = [r[0] for r in rows if len(r) == 1 and r[0]]
    assert "Plans" in section_titles
    assert "Groups" not in section_titles
    assert "Backup Servers" not in section_titles


def test_print_csv_details_sections(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, details = _make_detail_scenario()
    rows = _run_print_csv(
        plan_charges, group_charges, server_charges, capsys,
        details=details, pricing=make_two_plan_config(),
    )

    section_titles = [r[0] for r in rows if len(r) == 1]
    assert section_titles == [
        "Groups", "Backup Servers", "Plans", "Pricing Plans",
        "Group Backup Servers", "Group Plans", "Group Workloads",
        "Backup Servers by Workload Type", "Plans by Workload Type",
    ]
    pricing_start = rows.index(["Pricing Plans"])
    assert rows[pricing_start + 1] == ["name", "charge_per_instance", "charge_per_gb"]
    assert rows[pricing_start + 2] == ["Standard", "5.0", "0.2"]
    workloads_start = rows.index(["Group Workloads"])
    wl_header = rows[workloads_start + 1]
    wl_data = rows[workloads_start + 2]
    assert wl_data[wl_header.index("group_name")] == "Contoso"
    assert wl_data[wl_header.index("workload_type")] == "VM"
    assert wl_data[wl_header.index("storage_gb")] == "1.1084"
    assert wl_data[wl_header.index("total_charge")] == "20.33"


# ── table output ──────────────────────────────────────────────────────────────


@pytest.mark.parametrize("plan_type,expected", [
    ("Protection Plan", "Protection"),
    ("Retirement Plan", "Retirement"),
    ("Mixed", "Mixed"),
])
def test_kind_label(plan_type: str, expected: str) -> None:
    assert billing_report._kind_label(plan_type) == expected


def test_print_billing_table_rows_and_total(capsys: pytest.CaptureFixture[str]) -> None:
    charges = [make_plan_charge()]
    _print_billing_table(
        charges, [("Daily Backup", "Protection")], "Plan", "Plan Type",
        20, len("Pricing Plan"),
    )
    out = capsys.readouterr().out
    header = out.splitlines()[0]
    assert "Plan" in header
    assert "Instances" in header
    assert "Total Chg" in header
    row = _line_with(out, "Daily Backup")
    assert "Protection" in row
    assert "Standard" in row
    assert "2.00" in row  # storage_gb at 2 decimals
    assert "$15.00" in row
    assert "$0.40" in row
    assert "$15.40" in row
    total_row = [line for line in out.splitlines() if line.startswith("Total")][0]
    assert "$15.40" in total_row


def test_print_pricing_plans_table(capsys: pytest.CaptureFixture[str]) -> None:
    _print_pricing_plans_table(make_two_plan_config())
    out = capsys.readouterr().out
    assert out.splitlines()[0] == "Pricing Plans"
    standard = _line_with(out, "Standard")
    assert "$5.00" in standard
    assert "$0.20" in standard
    premium = _line_with(out, "Premium")
    assert "$10.00" in premium
    assert "$0.30" in premium


def test_print_pricing_plans_table_skips_unnamed_plans(capsys: pytest.CaptureFixture[str]) -> None:
    _print_pricing_plans_table(_PricingConfig(pricing_plans=[_PricingPlan("", 5.0, 0.1)]))
    assert capsys.readouterr().out == ""


def test_print_dist_table(capsys: pytest.CaptureFixture[str]) -> None:
    _print_dist_table(["Server"], [(["apm-server-01"], 2, 1.11)], widths=[13])
    out = capsys.readouterr().out
    header = out.splitlines()[0]
    assert "Server" in header
    assert "Instances" in header
    assert "Storage (GB)" in header
    row = _line_with(out, "apm-server-01")
    assert "        2" in row  # instances, right-aligned in a 9-wide column
    assert "1.11" in row


def test_print_details_tables_group_blocks_and_type_sections(
    capsys: pytest.CaptureFixture[str],
) -> None:
    plan_charges, group_charges, server_charges, details = _make_detail_scenario()
    _print_details_tables(group_charges, details, make_two_plan_config(), "")
    out = capsys.readouterr().out

    assert "Pricing Plans" in out
    assert "Group: Contoso  (Pricing Plan: Premium)" in out
    group_block = out[:out.index("Backup Servers by Workload Type")]
    workload_row = _line_with(group_block, "VM")
    assert "1.1084" in workload_row  # detail rows keep 4-decimal storage
    assert "$20.00" in workload_row
    assert "$0.33" in workload_row
    assert "$20.33" in workload_row
    assert "  Group Total: 2 instances, 1.11 GB, $20.33" in out.splitlines()
    server_section = out[out.index("Backup Servers by Workload Type"):]
    server_row = _line_with(server_section, "apm-server-01")
    assert "$10.22" in server_row
    plan_section = out[out.index("Plans by Workload Type"):]
    plan_row = _line_with(plan_section, "Daily Backup")
    assert "$10.22" in plan_row


def test_print_details_tables_only_servers(capsys: pytest.CaptureFixture[str]) -> None:
    _, group_charges, _, details = _make_detail_scenario()
    _print_details_tables(group_charges, details, make_two_plan_config(), "servers")
    out = capsys.readouterr().out
    assert "Backup Servers by Workload Type" in out
    assert "Group: Contoso" not in out
    assert "Plans by Workload Type" not in out


def test_print_table_no_workloads(capsys: pytest.CaptureFixture[str]) -> None:
    _print_table([], [], [], make_default_config(), None, "")
    assert capsys.readouterr().out.strip() == "(no workloads)"


def test_print_table_all_sections(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, _ = _make_detail_scenario()
    _print_table(plan_charges, group_charges, server_charges, make_two_plan_config(), None, "")
    out = capsys.readouterr().out

    group_row = _line_with(out, "Contoso")
    assert "Premium" in group_row
    assert "$20.00" in group_row
    assert "$0.33" in group_row
    assert "$20.33" in group_row
    server_row = _line_with(out, "apm-server-01")
    assert "$10.22" in server_row
    plan_row = _line_with(out, "Daily Backup")
    assert "Protection" in plan_row
    assert "$10.22" in plan_row


def test_print_table_only_plans_and_details_wiring(capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, details = _make_detail_scenario()
    _print_table(plan_charges, group_charges, server_charges, make_two_plan_config(), details, "plans")
    out = capsys.readouterr().out

    assert "Contoso" not in out
    assert "apm-server-01" not in out
    plan_table_row = _line_with(out.split("Plans by Workload Type")[0], "Daily Backup")
    assert "$10.22" in plan_table_row
    # details is not None → the per-type breakdown is appended
    assert "Plans by Workload Type" in out


# ── _write_xlsx ───────────────────────────────────────────────────────────────


def test_write_xlsx_summary_sheets(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, _ = _make_detail_scenario()
    path = tmp_path / "billing.xlsx"

    _write_xlsx(plan_charges, group_charges, server_charges, make_default_config(), None, "", str(path))

    saved_line = _line_with(capsys.readouterr().err, "Saved:")
    assert str(path) in saved_line
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Groups", "Backup Servers", "Plans"]
    groups = wb["Groups"]
    assert groups.cell(row=1, column=1).value == "group"
    assert groups.cell(row=2, column=1).value == "Contoso"
    assert groups.cell(row=2, column=3).value == "Premium"
    assert groups.cell(row=2, column=10).value == 20.33  # total_charge
    assert groups.cell(row=3, column=1).value == "Total"
    assert groups.cell(row=3, column=6).value == 2  # total instances
    plans = wb["Plans"]
    assert plans.cell(row=2, column=1).value == "Daily Backup"
    assert plans.cell(row=2, column=2).value == "plan-001"
    assert plans.cell(row=2, column=11).value == 10.22


def test_write_xlsx_detail_sheets(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    plan_charges, group_charges, server_charges, details = _make_detail_scenario()
    path = tmp_path / "billing.xlsx"

    _write_xlsx(
        plan_charges, group_charges, server_charges, make_two_plan_config(),
        details, "", str(path),
    )

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [
        "Groups", "Backup Servers", "Plans", "Pricing Plans",
        "Group Backup Servers", "Group Plans", "Group Workloads",
        "Backup Servers by Workload Type", "Plans by Workload Type",
    ]
    pricing_ws = wb["Pricing Plans"]
    assert pricing_ws.cell(row=2, column=1).value == "Standard"
    assert pricing_ws.cell(row=2, column=2).value == 5.0
    plan_type_ws = wb["Plans by Workload Type"]
    assert plan_type_ws.cell(row=1, column=5).value == "workload_type"
    assert plan_type_ws.cell(row=2, column=1).value == "Daily Backup"
    assert plan_type_ws.cell(row=2, column=5).value == "VM"
    assert plan_type_ws.cell(row=2, column=7).value == 1.1084
    assert plan_type_ws.cell(row=2, column=10).value == 10.22


def test_write_xlsx_empty_report_gets_placeholder_sheet(tmp_path: Path) -> None:
    path = tmp_path / "billing.xlsx"
    _write_xlsx([], [], [], make_default_config(), None, "", str(path))
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Report"]
