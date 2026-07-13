"""Unit tests for billing_report.py data collection and the run() entry point."""
from __future__ import annotations

import asyncio
import csv
import io
import json
from collections.abc import Callable, Coroutine
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import billing_report
import openpyxl
import pytest
from billing_report import (
    _bounded_paginate,
    _GroupSpec,
    _PricingConfig,
    _scan_billing,
    _sections_from_stats,
)

from synology_apm.sdk import (
    M365Workload,
    M365WorkloadType,
    MachineWorkload,
    MachineWorkloadType,
)
from tests.unit.examples._billing_fixtures import (
    STANDARD_RATE,
    make_default_config,
    make_server_stat,
)
from tests.unit.examples._fixtures import (
    make_backup_server,
    make_fake_apm,
    make_m365_workload,
    make_machine_workload,
    make_protection_plan,
    make_saas_tenant,
    patch_make_client,
)

_GB = 1024**3

# ── helpers ─────────────────────────────────────────────────────────────────────


def _machine_lister(
    active: list[MachineWorkload], retired: list[MachineWorkload]
) -> Callable[..., Coroutine[None, None, tuple[list[MachineWorkload], int]]]:
    """A machine workloads list() fake serving the active or retired set per is_retired."""

    async def _list(
        *, is_retired: bool, limit: int, offset: int
    ) -> tuple[list[MachineWorkload], int]:
        items = retired if is_retired else active
        return items[offset:offset + limit], len(items)

    return _list


def _m365_lister(
    items_by_key: dict[tuple[str, M365WorkloadType, bool], list[M365Workload]],
) -> Callable[..., Coroutine[None, None, tuple[list[M365Workload], int]]]:
    """An M365 workloads list() fake keyed by (tenant_id, workload_type, is_retired)."""

    async def _list(
        *, tenant_id: str, workload_type: M365WorkloadType, is_retired: bool, limit: int, offset: int
    ) -> tuple[list[M365Workload], int]:
        items = items_by_key.get((tenant_id, workload_type, is_retired), [])
        return items[offset:offset + limit], len(items)

    return _list


def _wire_machine(
    apm: MagicMock, active: list[MachineWorkload], retired: list[MachineWorkload] | None = None
) -> None:
    apm.machine.workloads.list = AsyncMock(side_effect=_machine_lister(active, retired or []))


def _line_with(out: str, needle: str) -> str:
    """The single output line containing *needle* (fails if absent or ambiguous)."""
    lines = [line for line in out.splitlines() if needle in line]
    assert len(lines) == 1, f"expected exactly one line containing {needle!r}, got {lines!r}"
    return lines[0]


# ── _bounded_paginate ──────────────────────────────────────────────────────────


async def test_bounded_paginate_drains_all_pages() -> None:
    wls = [
        make_machine_workload(workload_id=f"123e4567-e89b-12d3-a456-42661417400{i}")
        for i in range(3)
    ]

    async def list_call(limit: int, offset: int) -> tuple[list[MachineWorkload], int]:
        return wls[offset:offset + 2], len(wls)  # two-item pages

    items, total = await _bounded_paginate(asyncio.Semaphore(1), list_call)
    assert items == wls
    assert total == 3


async def test_bounded_paginate_waits_for_semaphore() -> None:
    sem = asyncio.Semaphore(1)
    await sem.acquire()

    async def list_call(limit: int, offset: int) -> tuple[list[MachineWorkload], int]:
        return [], 0

    task = asyncio.create_task(_bounded_paginate(sem, list_call))
    await asyncio.sleep(0)
    assert not task.done()  # blocked until the semaphore frees up
    sem.release()
    items, total = await task
    assert items == []
    assert total == 0


# ── _scan_billing ──────────────────────────────────────────────────────────────


async def test_scan_billing_tallies_per_server_plan_and_type() -> None:
    apm = make_fake_apm()
    plan = make_protection_plan(plan_id="plan-001", name="Daily Backup")
    workloads = [
        make_machine_workload(namespace="ns-a", plan=plan, workload_type=MachineWorkloadType.VM, protected_data_bytes=_GB),
        make_machine_workload(namespace="ns-a", plan=plan, workload_type=MachineWorkloadType.VM, protected_data_bytes=2 * _GB),
        make_machine_workload(namespace="ns-a", plan=plan, workload_type=MachineWorkloadType.PC, protected_data_bytes=_GB),
        make_machine_workload(namespace="ns-b", plan=plan, workload_type=MachineWorkloadType.VM, protected_data_bytes=_GB),
    ]
    _wire_machine(apm, workloads)

    sections, stats = await _scan_billing(apm, concurrency=2, pricing=make_default_config())

    assert len(stats) == 3
    by_key = {(s.namespace, s.type_label): s for s in stats}
    vm_a = by_key[("ns-a", "VM")]
    assert vm_a.plan_id == "plan-001"
    assert vm_a.group_names == ()
    assert vm_a.type_order == 0
    assert vm_a.count == 2
    assert vm_a.storage_bytes == 3 * _GB
    assert by_key[("ns-a", "PC")].count == 1
    assert by_key[("ns-a", "PC")].type_order == 2
    assert by_key[("ns-b", "VM")].count == 1
    # Sections aggregate the tallies over backup servers
    assert len(sections) == 1
    section = sections[0]
    assert section.plan_name == "Daily Backup"
    assert section.plan_type == "Protection Plan"
    assert [(r.type_label, r.count, r.storage_bytes) for r in section.rows] == [
        ("VM", 3, 4 * _GB),
        ("PC", 1, _GB),
    ]


async def test_scan_billing_classifies_retired_plans() -> None:
    apm = make_fake_apm()
    active_wl = make_machine_workload(plan=make_protection_plan(plan_id="plan-001", name="Daily Backup"))
    retired_wl = make_machine_workload(
        is_retired=True,
        plan=make_protection_plan(plan_id="plan-002", name="Compliance Retention"),
    )
    _wire_machine(apm, [active_wl], [retired_wl])

    sections, _ = await _scan_billing(apm, concurrency=2, pricing=make_default_config())

    assert {s.plan_name: s.plan_type for s in sections} == {
        "Daily Backup": "Protection Plan",
        "Compliance Retention": "Retirement Plan",
    }


async def test_scan_billing_first_seen_plan_meta_wins() -> None:
    # The same plan appears in both the active and the retired listing: the
    # active listing is processed first, so the plan stays a Protection Plan.
    apm = make_fake_apm()
    plan = make_protection_plan(plan_id="plan-001", name="Daily Backup")
    _wire_machine(apm, [make_machine_workload(plan=plan)], [make_machine_workload(plan=plan, is_retired=True)])

    sections, _ = await _scan_billing(apm, concurrency=2, pricing=make_default_config())

    assert len(sections) == 1
    assert sections[0].plan_type == "Protection Plan"
    assert sections[0].rows[0].count == 2  # both workloads still tallied


async def test_scan_billing_m365_fans_out_per_tenant_type_and_retired_flag() -> None:
    apm = make_fake_apm()
    tenant = make_saas_tenant()
    apm.saas.list = AsyncMock(return_value=([tenant], 1))
    exchange_wl = make_m365_workload(
        tenant_id=tenant.tenant_id, workload_type=M365WorkloadType.EXCHANGE,
        protected_data_bytes=_GB,
    )
    apm.m365.workloads.list = AsyncMock(side_effect=_m365_lister(
        {(tenant.tenant_id, M365WorkloadType.EXCHANGE, False): [exchange_wl]}
    ))

    sections, stats = await _scan_billing(apm, concurrency=3, pricing=make_default_config())

    calls = {
        (c.kwargs["tenant_id"], c.kwargs["workload_type"], c.kwargs["is_retired"])
        for c in apm.m365.workloads.list.call_args_list
    }
    assert calls == {
        (tenant.tenant_id, service, is_retired)
        for service in M365WorkloadType
        for is_retired in (False, True)
    }
    assert len(stats) == 1
    assert stats[0].type_label == "Exchange"
    assert stats[0].count == 1
    assert stats[0].storage_bytes == _GB
    assert len(sections) == 1
    assert sections[0].plan_type == "Protection Plan"


async def test_scan_billing_resolves_group_sets_by_plan() -> None:
    pricing = _PricingConfig(
        pricing_plans=[STANDARD_RATE],
        groups=[_GroupSpec("GroupA", "Standard", plan_ids=["plan-001"])],
    )
    apm = make_fake_apm()
    grouped = make_machine_workload(plan=make_protection_plan(plan_id="plan-001"))
    ungrouped = make_machine_workload(plan=make_protection_plan(plan_id="plan-002", name="Other Plan"))
    _wire_machine(apm, [grouped, ungrouped])

    sections, stats = await _scan_billing(apm, concurrency=1, pricing=pricing)

    assert {s.plan_id: s.group_names for s in stats} == {
        "plan-001": ("GroupA",),
        "plan-002": (),
    }
    assert {(s.plan_id, s.group_names) for s in sections} == {
        ("plan-001", ("GroupA",)),
        ("plan-002", ()),
    }


async def test_scan_billing_matches_groups_by_backup_server() -> None:
    pricing = _PricingConfig(
        pricing_plans=[STANDARD_RATE],
        groups=[_GroupSpec(
            "GroupB", "Standard", plan_ids=[],
            backup_server_ids=["123e4567-e89b-12d3-a456-426614174020"],
        )],
    )
    pricing.resolve_server_ids({"123e4567-e89b-12d3-a456-426614174020": "ns-apm-server-01"})
    apm = make_fake_apm()
    _wire_machine(apm, [make_machine_workload(namespace="ns-apm-server-01")])

    _, stats = await _scan_billing(apm, concurrency=1, pricing=pricing)

    assert len(stats) == 1
    assert stats[0].group_names == ("GroupB",)


# ── _sections_from_stats ───────────────────────────────────────────────────────


def test_sections_from_stats_same_plan_group_merges_counts() -> None:
    # Two stats for the same (plan_id, group_names) and same type_label → one row, summed
    s1 = make_server_stat(count=2, storage_bytes=_GB)
    s2 = make_server_stat(count=3, storage_bytes=2 * _GB)
    plan_meta: dict[str, tuple[str, str]] = {"plan-001": ("Daily Backup", "Protection Plan")}

    sections = _sections_from_stats([s1, s2], plan_meta)

    assert len(sections) == 1
    section = sections[0]
    assert section.plan_name == "Daily Backup"
    assert section.plan_type == "Protection Plan"
    assert len(section.rows) == 1
    assert section.rows[0].count == 5
    assert section.rows[0].storage_bytes == 3 * _GB


def test_sections_from_stats_different_group_names_produce_separate_sections() -> None:
    s1 = make_server_stat(group_names=("GroupA",), count=1)
    s2 = make_server_stat(group_names=("GroupB",), count=2)
    plan_meta: dict[str, tuple[str, str]] = {"plan-001": ("Daily Backup", "Protection Plan")}

    sections = _sections_from_stats([s1, s2], plan_meta)

    assert len(sections) == 2
    group_sets = {s.group_names for s in sections}
    assert ("GroupA",) in group_sets
    assert ("GroupB",) in group_sets


def test_sections_from_stats_sorted_by_plan_name_then_group_names() -> None:
    s_z = make_server_stat(plan_id="plan-z", count=1)
    s_a = make_server_stat(plan_id="plan-a", count=1)
    plan_meta: dict[str, tuple[str, str]] = {
        "plan-z": ("Zebra Plan", "Protection Plan"),
        "plan-a": ("Alpha Plan", "Protection Plan"),
    }

    sections = _sections_from_stats([s_z, s_a], plan_meta)

    assert sections[0].plan_name == "Alpha Plan"
    assert sections[1].plan_name == "Zebra Plan"


def test_sections_from_stats_rows_sorted_by_type_order() -> None:
    # Two stats with same (plan, group) but different type_order values
    s_high = make_server_stat(type_label="PC", type_order=2, count=1)
    s_low = make_server_stat(type_label="VM", type_order=0, count=1)
    plan_meta: dict[str, tuple[str, str]] = {"plan-001": ("Daily Backup", "Protection Plan")}

    sections = _sections_from_stats([s_high, s_low], plan_meta)

    assert len(sections) == 1
    rows = sections[0].rows
    assert len(rows) == 2
    assert rows[0].type_label == "VM"   # type_order=0 comes first
    assert rows[1].type_label == "PC"   # type_order=2 comes second


# ── run() ──────────────────────────────────────────────────────────────────────


def _wire_simple_scenario(apm: MagicMock) -> None:
    """One backup server (apm-server-01) with one 3-GB VM on Daily Backup (plan-001)."""
    apm.backup_servers.list = AsyncMock(return_value=([make_backup_server()], 1))
    plan = make_protection_plan(plan_id="plan-001", name="Daily Backup")
    wl = make_machine_workload(
        namespace="ns-apm-server-01", plan=plan,
        workload_type=MachineWorkloadType.VM, protected_data_bytes=3 * _GB,
    )
    _wire_machine(apm, [wl])


async def test_run_json_output(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    apm = make_fake_apm()
    _wire_simple_scenario(apm)
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="json", pricing=make_default_config(), show_details=False,
        concurrency=2, output_file=None, only="", configured_only=False,
    )

    captured = capsys.readouterr()
    assert "Collecting data..." in captured.err
    out = json.loads(captured.out)
    plan_row = out["plans"][0]
    assert plan_row["plan_name"] == "Daily Backup"
    assert plan_row["plan_type"] == "Protection Plan"
    assert plan_row["instances"] == 1
    assert plan_row["storage_gb"] == 3.0
    assert plan_row["instance_charge"] == 5.0
    assert plan_row["storage_charge"] == 0.60
    server_row = out["backup_servers"][0]
    assert server_row["server_name"] == "apm-server-01"
    assert server_row["namespace"] == "ns-apm-server-01"
    assert server_row["instances"] == 1


async def test_run_table_output(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    apm = make_fake_apm()
    _wire_simple_scenario(apm)
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="table", pricing=make_default_config(), show_details=False,
        concurrency=2, output_file=None, only="", configured_only=False,
    )

    out = capsys.readouterr().out
    plan_row = _line_with(out, "Daily Backup")
    assert "Protection" in plan_row
    assert "$5.00" in plan_row
    assert "$0.60" in plan_row
    assert "$5.60" in plan_row
    server_row = _line_with(out, "apm-server-01")
    assert "$5.60" in server_row


async def test_run_csv_output_with_details(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    apm = make_fake_apm()
    _wire_simple_scenario(apm)
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="csv", pricing=make_default_config(), show_details=True,
        concurrency=2, output_file=None, only="", configured_only=False,
    )

    rows = list(csv.reader(io.StringIO(capsys.readouterr().out)))
    section_titles = [r[0] for r in rows if len(r) == 1]
    # No groups configured → no Groups summary and no group detail sections
    assert section_titles == [
        "Backup Servers", "Plans", "Pricing Plans",
        "Backup Servers by Workload Type", "Plans by Workload Type",
    ]
    plans_header = rows[rows.index(["Plans"]) + 1]
    plans_data = rows[rows.index(["Plans"]) + 2]
    assert plans_data[plans_header.index("plan")] == "Daily Backup"
    assert plans_data[plans_header.index("storage_gb")] == "3.0"
    by_type_start = rows.index(["Plans by Workload Type"])
    by_type_header = rows[by_type_start + 1]
    by_type_data = rows[by_type_start + 2]
    assert by_type_data[by_type_header.index("workload_type")] == "VM"
    assert by_type_data[by_type_header.index("instance_charge")] == "5.0"


async def test_run_xlsx_output(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str], tmp_path: Path
) -> None:
    apm = make_fake_apm()
    _wire_simple_scenario(apm)
    patch_make_client(monkeypatch, billing_report, apm)
    output_file = tmp_path / "billing.xlsx"

    await billing_report.run(
        output_format="xlsx", pricing=make_default_config(), show_details=False,
        concurrency=2, output_file=str(output_file), only="", configured_only=False,
    )

    assert f"Saved: {output_file}" in capsys.readouterr().err
    wb = openpyxl.load_workbook(output_file)
    assert wb.sheetnames == ["Backup Servers", "Plans"]
    assert wb["Plans"].cell(row=2, column=1).value == "Daily Backup"


async def test_run_warns_on_unknown_backup_server_ids(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    pricing = _PricingConfig(
        pricing_plans=[STANDARD_RATE],
        groups=[_GroupSpec(
            "GroupA", "Standard", plan_ids=[],
            backup_server_ids=["123e4567-e89b-12d3-a456-426614174099"],
        )],
    )
    apm = make_fake_apm()
    _wire_simple_scenario(apm)  # the known server has a different backup_server_id
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="json", pricing=pricing, show_details=False,
        concurrency=2, output_file=None, only="", configured_only=False,
    )

    err = capsys.readouterr().err
    warning = _line_with(err, "unknown backup server ID(s)")
    assert "123e4567-e89b-12d3-a456-426614174099" in warning


async def test_run_warns_when_only_groups_matches_no_workloads(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    pricing = _PricingConfig(
        pricing_plans=[STANDARD_RATE],
        groups=[_GroupSpec("GroupA", "Standard", plan_ids=["plan-nomatch"])],
    )
    apm = make_fake_apm()
    _wire_simple_scenario(apm)  # workloads exist, but none on plan-nomatch
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="json", pricing=pricing, show_details=False,
        concurrency=2, output_file=None, only="groups", configured_only=False,
    )

    captured = capsys.readouterr()
    assert "no workloads matched any group" in captured.err
    assert json.loads(captured.out)["groups"] == []


async def test_run_configured_only_filters_plans_and_servers(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    pricing = _PricingConfig(
        pricing_plans=[STANDARD_RATE],
        assignments={"plan-001": "Standard"},
    )
    apm = make_fake_apm()
    apm.backup_servers.list = AsyncMock(return_value=([make_backup_server()], 1))
    configured = make_machine_workload(plan=make_protection_plan(plan_id="plan-001", name="Daily Backup"))
    unconfigured = make_machine_workload(plan=make_protection_plan(plan_id="plan-002", name="Other Plan"))
    _wire_machine(apm, [configured, unconfigured])
    patch_make_client(monkeypatch, billing_report, apm)

    await billing_report.run(
        output_format="json", pricing=pricing, show_details=False,
        concurrency=2, output_file=None, only="", configured_only=True,
    )

    out = json.loads(capsys.readouterr().out)
    assert [p["plan_name"] for p in out["plans"]] == ["Daily Backup"]
    # No backup server is mentioned in the config → the servers section is filtered empty
    assert out["backup_servers"] == []
