#!/usr/bin/env python3
"""
Billing report — computes workload count, storage usage, and charges across three
independent billing dimensions: groups, backup servers, and APM plans
(Protection Plans and Retirement Plans, covering both machine and M365 workloads).

The pricing config (--config) has four layers:
  - pricing_plans: named rate cards (charge per instance, charge per GB); the first
    entry is the fallback rate for anything not explicitly assigned
  - groups: named workload groupings — each aggregates its workloads into a single
    billing row using the group's pricing plan. Membership is by APM plan (plans) and/or
    by backup server ID (backup_servers): a workload belongs to every group that lists
    its plan or its backup server, so groups may overlap.
  - backup_servers: per-server pricing for the Backup Servers section
  - plans: per-plan pricing for the Plans section

Output is split into three sections — Groups (one row per group), Backup Servers
(one row per server), and Plans (one row per plan). The sections are independent
views over the same workloads: Backup Servers and Plans each count every workload
exactly once, and a grouped workload additionally appears in the row of every group
it belongs to. Pass --only groups|servers|plans to print a single section. When
--config is given, the Backup Servers and Plans sections list only servers/plans
mentioned in the config; pass --show-not-configured to include the rest.

Pass --charge-per-instance / --charge-per-gb to apply a single flat rate to all
servers and plans without a YAML config. Pass --dump-config-template to generate a
starter template and exit (no APM connection needed).

Pass --details for a per-workload-type breakdown: one block per group showing its
backup servers and plans as charge-free workload distributions plus a charged
per-workload-type Workloads table (with a group total line), followed by
"Backup Servers by Workload Type" and "Plans by Workload Type" sections. The rate
cards are also listed as a Pricing Plans section. CSV and XLSX mirror the group
blocks with flat sections/sheets; JSON nests the distributions and workloads under
each group and a by_type list under each server and plan. Each detail view rounds
its charges independently (per workload type, per plan, per backup server), so
summing one view's displayed rows may differ from another view's printed total by a
few cents.

Usage:
    python billing_report.py
    python billing_report.py --charge-per-instance 5 --charge-per-gb 0.1
    python billing_report.py --config pricing.yaml
    python billing_report.py --config pricing.yaml --details
    python billing_report.py --dump-config-template > pricing.yaml
    python billing_report.py -o xlsx --output-file billing.xlsx

Environment variables (see .env.example and examples/README.md):
    APM_HOST          hostname or IP (supports host:port)
    APM_USERNAME      account
    APM_PASSWORD      password
    APM_NO_VERIFY_SSL set to true to skip SSL verification (self-signed certificate environments)

Additional dependencies (dev deps, included by `uv sync`):
    pip install pyyaml    # for --config / --dump-config-template
    pip install openpyxl  # for -o xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
from collections import Counter
from collections.abc import Awaitable, Callable, Sequence
from dataclasses import asdict, dataclass, field
from typing import Any, NamedTuple, TypeVar

import openpyxl
import openpyxl.styles
import yaml
from _common import (
    WORKLOAD_TYPE_ORDER,
    add_profile_arg,
    collect_backup_servers,
    list_m365_tenants,
    make_client,
    paginate,
    run_main,
    workload_type_label,
)
from openpyxl.utils import get_column_letter as _get_col_letter
from openpyxl.worksheet.worksheet import Worksheet as _Worksheet

from synology_apm.sdk import (
    APMClient,
    M365Workload,
    M365WorkloadType,
    MachineWorkload,
    SaasTenant,
)

_PROTECTION = "Protection Plan"
_RETIREMENT = "Retirement Plan"
_MIXED_KIND = "Mixed"
_BYTES_PER_GB = 1024 ** 3
_DEFAULT_CONCURRENCY = 5
_UNKNOWN_SERVER = "(unknown)"

_M365_TYPES: list[M365WorkloadType] = [t for t in WORKLOAD_TYPE_ORDER if isinstance(t, M365WorkloadType)]

_TYPE_ORDER_INDEX = {t: i for i, t in enumerate(WORKLOAD_TYPE_ORDER)}


# ── Pricing config ────────────────────────────────────────────────────────────

@dataclass
class _PricingPlan:
    name: str
    charge_per_instance: float = 0.0
    charge_per_gb: float = 0.0


@dataclass
class _GroupSpec:
    name: str
    pricing_plan_name: str
    plan_ids: list[str]
    backup_server_ids: list[str] = field(default_factory=list)


@dataclass
class _PricingConfig:
    """Named pricing plans with groups and per-dimension assignments.

    Rate resolution is per dimension: pricing_plan_for_plan() for the Plans section
    (plan assignment or fallback), pricing_plan_for_server() for the Backup Servers
    section (server assignment or fallback), and pricing_plan_for_group() for group
    rows. The first pricing plan is the fallback for anything unassigned.

    A workload belongs to every group that lists its backup server or its plan —
    groups may overlap, and a workload matching one group through both a backup
    server and a plan is still that group's member once. groups_for() resolves the
    group set per workload.

    Backup servers are configured by backup server ID, but workloads are matched by
    namespace: resolve_server_ids() must be called once with the ID → namespace map
    (from the server list) before groups_for() / pricing_plan_for_server() can match
    by backup server.

    Pricing plan and group names are assumed unique (enforced by _load_pricing_yaml).
    """
    pricing_plans: list[_PricingPlan]
    groups: list[_GroupSpec] = field(default_factory=list)
    assignments: dict[str, str] = field(default_factory=dict)
    server_assignments: dict[str, str] = field(default_factory=dict)
    _pp_by_name: dict[str, _PricingPlan] = field(init=False, default_factory=dict)
    _plan_to_groups: dict[str, list[str]] = field(init=False, default_factory=dict)
    _group_pricing: dict[str, _PricingPlan] = field(init=False, default_factory=dict)
    _group_order: dict[str, int] = field(init=False, default_factory=dict)
    _ns_to_groups: dict[str, list[str]] = field(init=False, default_factory=dict)
    _ns_assignments: dict[str, str] = field(init=False, default_factory=dict)

    def __post_init__(self) -> None:
        self._pp_by_name = {pp.name: pp for pp in self.pricing_plans}
        for idx, group in enumerate(self.groups):
            self._group_order[group.name] = idx
            self._group_pricing[group.name] = self._pp_by_name.get(
                group.pricing_plan_name, self.pricing_plans[0]
            )
            for pid in group.plan_ids:
                self._plan_to_groups.setdefault(pid, []).append(group.name)

    def resolve_server_ids(self, ns_by_server_id: dict[str, str]) -> list[str]:
        """Translate configured backup server IDs into workload namespaces.

        Must be called once, after fetching the backup server list, before
        groups_for() / pricing_plan_for_server() can match by backup server.
        Returns the configured IDs that match no known backup server, sorted.
        """
        unknown: set[str] = set()
        for group in self.groups:
            for sid in group.backup_server_ids:
                ns = ns_by_server_id.get(sid)
                if ns is None:
                    unknown.add(sid)
                else:
                    self._ns_to_groups.setdefault(ns, []).append(group.name)
        for sid, pp_name in self.server_assignments.items():
            ns = ns_by_server_id.get(sid)
            if ns is None:
                unknown.add(sid)
            else:
                self._ns_assignments[ns] = pp_name
        return sorted(unknown)

    def groups_for(self, plan_id: str, namespace: str) -> tuple[str, ...]:
        """Return the names of every group one workload belongs to, in config order.

        Returns () for an ungrouped workload. A workload matching one group through
        both a backup server and a plan appears once.
        """
        matched = set(self._ns_to_groups.get(namespace, ()) if namespace else ())
        matched.update(self._plan_to_groups.get(plan_id, ()))
        return tuple(sorted(matched, key=self._group_order.__getitem__))

    def pricing_plan_for_plan(self, plan_id: str) -> _PricingPlan:
        """Return the rate card for one plan row of the Plans section."""
        name = self.assignments.get(plan_id)
        if name:
            return self._pp_by_name.get(name, self.pricing_plans[0])
        return self.pricing_plans[0]

    def pricing_plan_for_server(self, namespace: str) -> _PricingPlan:
        """Return the rate card for one server row of the Backup Servers section."""
        name = self._ns_assignments.get(namespace)
        if name:
            return self._pp_by_name.get(name, self.pricing_plans[0])
        return self.pricing_plans[0]

    def pricing_plan_for_group(self, group_name: str) -> _PricingPlan:
        """Return the rate card for one group row of the Groups section."""
        return self._group_pricing.get(group_name, self.pricing_plans[0])

    @property
    def configured_plan_ids(self) -> set[str]:
        """Plan IDs mentioned anywhere in the config (rate assignment or group member)."""
        return set(self.assignments) | set(self._plan_to_groups)

    @property
    def configured_namespaces(self) -> set[str]:
        """Namespaces of backup servers mentioned anywhere in the config (resolved)."""
        return set(self._ns_assignments) | set(self._ns_to_groups)


# ── Data models ───────────────────────────────────────────────────────────────

@dataclass
class _TypeCount:
    type_label: str
    type_order: int
    count: int
    storage_bytes: int


@dataclass
class _PlanSection:
    """Per-(plan, group set) workload tallies, aggregated over backup servers.

    group_names is the full set of groups the section's workloads belong to
    (config order; empty for ungrouped workloads), so each workload appears in
    exactly one section.
    """
    plan_name: str
    plan_type: str
    plan_id: str
    group_names: tuple[str, ...]
    rows: list[_TypeCount] = field(default_factory=list)

    @property
    def total_count(self) -> int:
        return sum(r.count for r in self.rows)

    @property
    def total_bytes(self) -> int:
        return sum(r.storage_bytes for r in self.rows)


@dataclass(kw_only=True)
class _ChargedRow:
    """Numeric charge fields and the derived total shared by every charged row."""
    instances: int
    storage_gb: float
    instance_charge: float
    storage_charge: float

    @property
    def total_charge(self) -> float:
        return self.instance_charge + self.storage_charge


@dataclass(kw_only=True)
class _RatedCharge(_ChargedRow):
    """A summary charge row carrying its resolved rate card (Groups/Servers/Plans)."""
    pricing_plan_name: str
    charge_per_instance: float
    charge_per_gb: float


@dataclass(kw_only=True)
class _TypeChargeRow(_ChargedRow):
    """A per-workload-type detail row (rate-card name + type label)."""
    pricing_plan: str
    workload_type: str


@dataclass(kw_only=True)
class _PlanCharge(_RatedCharge):
    plan_name: str
    plan_type: str
    plan_id: str


@dataclass(kw_only=True)
class _GroupCharge(_RatedCharge):
    group_name: str
    plan_type: str


@dataclass(kw_only=True)
class _ServerCharge(_RatedCharge):
    server_name: str
    namespace: str


@dataclass
class _ServerTypeStat:
    """Per-(server, plan, group set, workload type) tallies — the single granular base
    of all three dimensions; rate cards are resolved later. Each workload increments
    exactly one stat: group_names carries every group it belongs to.
    """
    namespace: str
    plan_id: str
    group_names: tuple[str, ...]
    type_label: str
    type_order: int
    count: int
    storage_bytes: int


@dataclass(kw_only=True)
class _DetailRow(_TypeChargeRow):
    """One (plan, workload type) row of the Plans by Workload Type breakdown."""
    plan_type: str
    plan_name: str
    plan_id: str


@dataclass(kw_only=True)
class _ServerDetailRow(_TypeChargeRow):
    """One (server, workload type) row of the Backup Servers by Workload Type breakdown."""
    server_name: str
    namespace: str


@dataclass
class _GroupServerRow:
    """One server's workload distribution inside a group block (no charges)."""
    group_name: str
    server_name: str
    namespace: str
    instances: int
    storage_gb: float


@dataclass
class _GroupPlanRow:
    """One plan's workload distribution inside a group block (no charges)."""
    group_name: str
    plan_name: str
    plan_id: str
    plan_type: str
    instances: int
    storage_gb: float


@dataclass(kw_only=True)
class _GroupTypeRow(_TypeChargeRow):
    """One (group, workload type) row, charged at the group's rate card."""
    group_name: str


@dataclass
class _DetailsView:
    """Detail rows for every output format.

    The by-group dicts are keyed in group-charge (YAML) order so all formats render
    groups in the same order; row order inside each list is set by the builders.
    """
    server_rows_by_group: dict[str, list[_GroupServerRow]]
    plan_rows_by_group: dict[str, list[_GroupPlanRow]]
    workload_rows_by_group: dict[str, list[_GroupTypeRow]]
    server_type_rows: list[_ServerDetailRow]
    plan_type_rows: list[_DetailRow]


def fmt_money(amount: float) -> str:
    """Render a charge amount with 2 decimal places and thousands separators."""
    return f"${amount:,.2f}"


def _server_display_name(server_names: dict[str, str], namespace: str) -> str:
    """Display name for a backup server, falling back to its namespace or a placeholder."""
    return server_names.get(namespace) or namespace or _UNKNOWN_SERVER


# ── Pricing config I/O ────────────────────────────────────────────────────────

_PRICING_TEMPLATE = """\
# APM billing report pricing configuration
# Generate this template: python billing_report.py --dump-config-template > pricing.yaml
#
# The report bills three independent dimensions: Groups, Backup Servers, and
# Plans. A workload is counted in every section it matches.
# By default only the servers/plans mentioned in this file are listed; pass
# --show-not-configured to also list the rest, charged at the fallback rate.

# Rate cards to assign to groups, backup servers, or APM plans.
# The first entry is the fallback for anything not explicitly assigned.
# Both charge fields default to 0 when omitted.
pricing_plans:
  - name: Standard
    charge_per_instance: 5.0   # charge per workload instance
    charge_per_gb: 0.10        # charge per GB of protected storage
  - name: Premium
    charge_per_instance: 10.0
    charge_per_gb: 0.20
  - name: Compliance
    charge_per_instance: 0.0
    charge_per_gb: 0.05

# Each group is billed as one combined row. A workload belongs to a group when
# it is on one of the group's backup servers (by backup server ID)
# OR its APM plan (Protection Plan or Retirement Plan) is listed in plans.
# Either membership list may be omitted. Groups may overlap: a server/plan
# listed in several groups puts its workloads in each of those groups.
# Find plan IDs: python billing_report.py -o json (the plan_id field in plans).
# Find backup server IDs: synology-apm-cli infra server list --verbose, or
# BackupServer.backup_server_id from apm.backup_servers.list().
groups:
  - name: Contoso
    pricing_plan: Premium
    plans:
      - 123e4567-e89b-12d3-a456-426614174001  # Daily Backup
      - 123e4567-e89b-12d3-a456-426614174002  # Compliance Retention
    backup_servers:
      - 123e4567-e89b-12d3-a456-426614174010  # apm-server-01

# Optional per-server rates for the Backup Servers section: backup server ID → pricing plan.
backup_servers:
  # 123e4567-e89b-12d3-a456-426614174011: Premium

# Optional per-plan rates for the Plans section: plan ID → pricing plan.
plans:
  # 123e4567-e89b-12d3-a456-426614174003: Compliance
"""


def _dump_config_template() -> None:
    print(_PRICING_TEMPLATE, end="")


def _parse_membership_list(path: str, gname: str, field_name: str, raw: Any) -> list[str]:
    """Parse one group membership list (plans / backup_servers): unique non-blank strings."""
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError(f"{path}: group '{gname}': {field_name} must be a list")
    if any(item is None or not str(item).strip() for item in raw):
        raise ValueError(f"{path}: group '{gname}' contains a blank entry in {field_name}")
    values = [str(item).strip() for item in raw]
    duplicates = sorted(v for v, n in Counter(values).items() if n > 1)
    if duplicates:
        raise ValueError(
            f"{path}: group '{gname}': duplicate entries in {field_name}: {', '.join(duplicates)}"
        )
    return values


def _parse_assignments(path: str, key: str, raw: Any, pricing_plan_names: set[str]) -> dict[str, str]:
    """Parse one top-level assignment dict (plans / backup_servers): ID → rate-card name."""
    assignments = {str(k): str(v) for k, v in (raw or {}).items()}
    unknown = set(assignments.values()) - pricing_plan_names
    if unknown:
        raise ValueError(f"{path}: unknown pricing plan name(s) in {key}: {', '.join(sorted(unknown))}")
    return assignments


def _load_pricing_yaml(path: str) -> _PricingConfig:
    with open(path) as f:
        raw: dict[str, Any] = yaml.safe_load(f)
    pricing_plans = [
        _PricingPlan(
            name=str(entry["name"]),
            charge_per_instance=float(entry.get("charge_per_instance", 0.0)),
            charge_per_gb=float(entry.get("charge_per_gb", 0.0)),
        )
        for entry in (raw.get("pricing_plans") or [])
    ]
    if not pricing_plans:
        raise ValueError(f"{path}: pricing_plans must contain at least one entry")
    names = [pp.name for pp in pricing_plans]
    duplicate_names = sorted(n for n, cnt in Counter(names).items() if cnt > 1)
    if duplicate_names:
        raise ValueError(
            f"{path}: duplicate pricing plan name(s): {', '.join(duplicate_names)}"
        )
    pricing_plan_names = set(names)

    groups: list[_GroupSpec] = []
    seen_group_names: set[str] = set()
    for entry in (raw.get("groups") or []):
        gname = str(entry["name"])
        if not gname.strip():
            raise ValueError(f"{path}: group names must not be blank")
        if gname in seen_group_names:
            raise ValueError(f"{path}: duplicate group name '{gname}'")
        seen_group_names.add(gname)
        pp_name = str(entry["pricing_plan"])
        if pp_name not in pricing_plan_names:
            raise ValueError(f"{path}: group '{gname}' references unknown pricing plan '{pp_name}'")
        plan_ids = _parse_membership_list(path, gname, "plans", entry.get("plans"))
        server_ids = _parse_membership_list(path, gname, "backup_servers", entry.get("backup_servers"))
        groups.append(_GroupSpec(
            name=gname, pricing_plan_name=pp_name, plan_ids=plan_ids, backup_server_ids=server_ids,
        ))

    return _PricingConfig(
        pricing_plans=pricing_plans,
        groups=groups,
        assignments=_parse_assignments(path, "plans", raw.get("plans"), pricing_plan_names),
        server_assignments=_parse_assignments(
            path, "backup_servers", raw.get("backup_servers"), pricing_plan_names),
    )


# ── Data collection ──────────────────────────────────────────────────────────

def _sections_from_stats(
    server_stats: list[_ServerTypeStat], plan_meta: dict[str, tuple[str, str]],
) -> list[_PlanSection]:
    """One _PlanSection per (plan, group set): stats summed over backup servers.

    Sections are sorted by plan name (then group set); within a section, rows follow
    the canonical workload-type display order.
    """
    buckets: dict[tuple[str, tuple[str, ...]], dict[str, _TypeCount]] = {}
    for st in server_stats:
        per_type = buckets.setdefault((st.plan_id, st.group_names), {})
        tc = per_type.get(st.type_label)
        if tc is None:
            tc = per_type[st.type_label] = _TypeCount(st.type_label, st.type_order, 0, 0)
        tc.count += st.count
        tc.storage_bytes += st.storage_bytes
    sections = []
    for (pid, group_names), per_type in buckets.items():
        plan_name, plan_type = plan_meta[pid]
        rows = sorted(per_type.values(), key=lambda r: r.type_order)
        sections.append(_PlanSection(plan_name, plan_type, pid, group_names, rows))
    sections.sort(key=lambda s: (s.plan_name, s.group_names))
    return sections


_WorkloadT = TypeVar("_WorkloadT", MachineWorkload, M365Workload)


async def _bounded_paginate(
    sem: asyncio.Semaphore,
    list_call: Callable[[int, int], Awaitable[tuple[list[_WorkloadT], int | None]]],
) -> tuple[list[_WorkloadT], int | None]:
    """Run one paginate() call under *sem*, so --concurrency bounds concurrent API calls directly."""
    async with sem:
        return await paginate(list_call)


async def _scan_billing(
    apm: APMClient, *, concurrency: int, pricing: _PricingConfig
) -> tuple[list[_PlanSection], list[_ServerTypeStat]]:
    """One _PlanSection per (plan, group set) pair with at least one workload.

    Fetches all workloads in a single pass (machine active/retired + M365 per tenant/type
    active/retired) and tallies them per (backup server, plan, group set, workload type)
    — the single granular base of all three dimensions and the --details breakdowns.
    Each workload lands in exactly one tally: the plan comes from the lightweight plan
    reference embedded in each workload, the group set from pricing.groups_for() (which
    requires pricing.resolve_server_ids() to have been called). Sections aggregate the
    tallies over backup servers; a plan whose workloads span different group sets
    produces multiple sections. Section and row order: see _sections_from_stats.
    """
    sem = asyncio.Semaphore(concurrency)
    tenants = await list_m365_tenants(apm)

    # (is_retired, paginate coroutine) pairs; the flag classifies each result's plan kind.
    tasks: list[tuple[bool, Awaitable[tuple[list[MachineWorkload] | list[M365Workload], int | None]]]] = []

    for is_retired in (False, True):
        async def _machine(
            limit: int, offset: int, r: bool = is_retired
        ) -> tuple[list[MachineWorkload], int | None]:
            return await apm.machine.workloads.list(is_retired=r, limit=limit, offset=offset)
        tasks.append((is_retired, _bounded_paginate(sem, _machine)))

    for tenant in tenants:
        for service in _M365_TYPES:
            for is_retired in (False, True):
                async def _m365(
                    limit: int, offset: int,
                    t: SaasTenant = tenant, s: M365WorkloadType = service, r: bool = is_retired,
                ) -> tuple[list[M365Workload], int | None]:
                    return await apm.m365.workloads.list(
                        tenant_id=t.tenant_id, workload_type=s, is_retired=r, limit=limit, offset=offset,
                    )
                tasks.append((is_retired, _bounded_paginate(sem, _m365)))

    plan_meta: dict[str, tuple[str, str]] = {}
    server_stats: dict[tuple[str, str, tuple[str, ...], str], _ServerTypeStat] = {}

    results = await asyncio.gather(*(coro for _, coro in tasks))
    for (is_retired, _), (items, _) in zip(tasks, results, strict=True):
        for wl in items:
            pid = wl.plan.plan_id
            if pid not in plan_meta:
                plan_meta[pid] = (wl.plan.name, _RETIREMENT if is_retired else _PROTECTION)
            group_names = pricing.groups_for(pid, wl.namespace)

            type_label = workload_type_label(wl)
            stat_key = (wl.namespace, pid, group_names, type_label)
            stat = server_stats.get(stat_key)
            if stat is None:
                stat = server_stats[stat_key] = _ServerTypeStat(
                    namespace=wl.namespace,
                    plan_id=pid,
                    group_names=group_names,
                    type_label=type_label,
                    type_order=_TYPE_ORDER_INDEX[wl.workload_type],
                    count=0,
                    storage_bytes=0,
                )
            stat.count += 1
            stat.storage_bytes += wl.protected_data_bytes

    stats = list(server_stats.values())
    return _sections_from_stats(stats, plan_meta), stats


# ── Charge computation ────────────────────────────────────────────────────────

@dataclass
class _Tally:
    """Mutable (workload count, protected bytes) accumulator."""
    count: int = 0
    storage_bytes: int = 0

    def add(self, count: int, storage_bytes: int) -> None:
        self.count += count
        self.storage_bytes += storage_bytes


def _charge_fields(
    count: int, storage_bytes: int, charge_per_instance: float, charge_per_gb: float,
    *, gb_decimals: int,
) -> tuple[float, float, float]:
    """(rounded storage_gb, instance_charge, storage_charge) for one charged row.

    Charges are computed from the unrounded GB value; only the returned
    storage_gb is rounded (2 decimals for summary rows, 4 for detail rows).
    """
    storage_gb = storage_bytes / _BYTES_PER_GB
    return (
        round(storage_gb, gb_decimals),
        round(count * charge_per_instance, 2),
        round(storage_gb * charge_per_gb, 2),
    )


def _compute_plan_charges(sections: list[_PlanSection], pricing: _PricingConfig) -> list[_PlanCharge]:
    """One charge row per plan, merged across groups, sorted by plan name."""
    merged: dict[str, list[_PlanSection]] = {}
    for s in sections:
        merged.setdefault(s.plan_id, []).append(s)
    charges = []
    for pid, secs in merged.items():
        pp = pricing.pricing_plan_for_plan(pid)
        count = sum(s.total_count for s in secs)
        storage_gb, instance_charge, storage_charge = _charge_fields(
            count, sum(s.total_bytes for s in secs),
            pp.charge_per_instance, pp.charge_per_gb, gb_decimals=2,
        )
        charges.append(_PlanCharge(
            plan_name=secs[0].plan_name,
            plan_type=secs[0].plan_type,
            plan_id=pid,
            instances=count,
            storage_gb=storage_gb,
            pricing_plan_name=pp.name,
            charge_per_instance=pp.charge_per_instance,
            charge_per_gb=pp.charge_per_gb,
            instance_charge=instance_charge,
            storage_charge=storage_charge,
        ))
    charges.sort(key=lambda c: (c.plan_name, c.plan_id))
    return charges


def _compute_group_charges(sections: list[_PlanSection], pricing: _PricingConfig) -> list[_GroupCharge]:
    """One charge row per group with at least one workload, in YAML group order.

    A section belongs to every group in its group set, so overlapping groups each
    count the shared workloads (and the section's Total repeats them per group).
    """
    buckets: dict[str, list[_PlanSection]] = {g.name: [] for g in pricing.groups}
    for s in sections:
        for gname in s.group_names:
            if gname in buckets:
                buckets[gname].append(s)
    result = []
    for g in pricing.groups:
        members = buckets[g.name]
        if not members:
            continue
        kinds = {m.plan_type for m in members}
        plan_type = kinds.pop() if len(kinds) == 1 else _MIXED_KIND
        pp = pricing.pricing_plan_for_group(g.name)
        count = sum(s.total_count for s in members)
        storage_gb, instance_charge, storage_charge = _charge_fields(
            count, sum(s.total_bytes for s in members),
            pp.charge_per_instance, pp.charge_per_gb, gb_decimals=2,
        )
        result.append(_GroupCharge(
            group_name=g.name,
            plan_type=plan_type,
            pricing_plan_name=pp.name,
            charge_per_instance=pp.charge_per_instance,
            charge_per_gb=pp.charge_per_gb,
            instances=count,
            storage_gb=storage_gb,
            instance_charge=instance_charge,
            storage_charge=storage_charge,
        ))
    return result


def _compute_server_charges(
    server_stats: list[_ServerTypeStat],
    pricing: _PricingConfig,
    server_names: dict[str, str],
) -> list[_ServerCharge]:
    """One charge row per backup server over all its workloads, sorted by server name."""
    buckets: dict[str, _Tally] = {}
    for st in server_stats:
        buckets.setdefault(st.namespace, _Tally()).add(st.count, st.storage_bytes)
    charges = []
    for ns, tally in buckets.items():
        pp = pricing.pricing_plan_for_server(ns)
        storage_gb, instance_charge, storage_charge = _charge_fields(
            tally.count, tally.storage_bytes,
            pp.charge_per_instance, pp.charge_per_gb, gb_decimals=2,
        )
        charges.append(_ServerCharge(
            server_name=_server_display_name(server_names, ns),
            namespace=ns,
            pricing_plan_name=pp.name,
            charge_per_instance=pp.charge_per_instance,
            charge_per_gb=pp.charge_per_gb,
            instances=tally.count,
            storage_gb=storage_gb,
            instance_charge=instance_charge,
            storage_charge=storage_charge,
        ))
    charges.sort(key=lambda c: (c.server_name, c.namespace))
    return charges


@dataclass
class _ChargeTotals:
    instances: int
    storage_gb: float
    instance_charge: float
    storage_charge: float
    total_charge: float


def _aggregate(charges: Sequence[_ChargedRow]) -> _ChargeTotals:
    """Five-field totals over charge rows of any dimension, rounded for display."""
    return _ChargeTotals(
        instances=sum(c.instances for c in charges),
        storage_gb=round(sum(c.storage_gb for c in charges), 2),
        instance_charge=round(sum(c.instance_charge for c in charges), 2),
        storage_charge=round(sum(c.storage_charge for c in charges), 2),
        total_charge=round(sum(c.total_charge for c in charges), 2),
    )


# ── Details data ─────────────────────────────────────────────────────────────

def _build_group_server_rows(
    server_stats: list[_ServerTypeStat],
    group_charges: list[_GroupCharge],
    server_names: dict[str, str],
) -> dict[str, list[_GroupServerRow]]:
    """Per group: one distribution row per backup server, sorted by server name."""
    buckets: dict[tuple[str, str], _Tally] = {}
    for st in server_stats:
        for gname in st.group_names:
            buckets.setdefault((gname, st.namespace), _Tally()).add(st.count, st.storage_bytes)
    rows_by_group: dict[str, list[_GroupServerRow]] = {g.group_name: [] for g in group_charges}
    for (gname, ns), tally in buckets.items():
        if gname not in rows_by_group:
            continue
        rows_by_group[gname].append(_GroupServerRow(
            group_name=gname,
            server_name=_server_display_name(server_names, ns),
            namespace=ns,
            instances=tally.count,
            storage_gb=round(tally.storage_bytes / _BYTES_PER_GB, 2),
        ))
    for rows in rows_by_group.values():
        rows.sort(key=lambda r: (r.server_name, r.namespace))
    return rows_by_group


def _build_group_plan_rows(
    sections: list[_PlanSection], group_charges: list[_GroupCharge],
) -> dict[str, list[_GroupPlanRow]]:
    """Per group: one distribution row per plan, in section (plan-name) order.

    A plan's workloads can span several sections visible to the same group (one per
    group set), so rows are merged by (group, plan).
    """
    rows_by_group: dict[str, list[_GroupPlanRow]] = {g.group_name: [] for g in group_charges}
    buckets: dict[tuple[str, str], _Tally] = {}
    meta: dict[tuple[str, str], tuple[str, str]] = {}
    for s in sections:
        for gname in s.group_names:
            if gname not in rows_by_group:
                continue
            key = (gname, s.plan_id)
            b = buckets.get(key)
            if b is None:
                b = buckets[key] = _Tally()
                meta[key] = (s.plan_name, s.plan_type)
            b.add(s.total_count, s.total_bytes)
    for (gname, pid), tally in buckets.items():
        plan_name, plan_type = meta[(gname, pid)]
        rows_by_group[gname].append(_GroupPlanRow(
            group_name=gname,
            plan_name=plan_name,
            plan_id=pid,
            plan_type=plan_type,
            instances=tally.count,
            storage_gb=round(tally.storage_bytes / _BYTES_PER_GB, 2),
        ))
    return rows_by_group


def _build_group_workload_rows(
    sections: list[_PlanSection], group_charges: list[_GroupCharge],
) -> dict[str, list[_GroupTypeRow]]:
    """Per group: one charged row per workload type, in canonical type order."""
    secs_by_group: dict[str, list[_PlanSection]] = {}
    for s in sections:
        for gname in s.group_names:
            secs_by_group.setdefault(gname, []).append(s)
    rows_by_group: dict[str, list[_GroupTypeRow]] = {}
    for g in group_charges:
        buckets: dict[str, _Tally] = {}
        order: dict[str, int] = {}
        for s in secs_by_group.get(g.group_name, []):
            for tr in s.rows:
                if tr.type_label not in buckets:
                    buckets[tr.type_label] = _Tally()
                    order[tr.type_label] = tr.type_order
                buckets[tr.type_label].add(tr.count, tr.storage_bytes)
        rows = []
        for label, tally in buckets.items():
            storage_gb, instance_charge, storage_charge = _charge_fields(
                tally.count, tally.storage_bytes,
                g.charge_per_instance, g.charge_per_gb, gb_decimals=4,
            )
            rows.append(_GroupTypeRow(
                group_name=g.group_name,
                pricing_plan=g.pricing_plan_name,
                workload_type=label,
                instances=tally.count,
                storage_gb=storage_gb,
                instance_charge=instance_charge,
                storage_charge=storage_charge,
            ))
        rows.sort(key=lambda r: order[r.workload_type])
        rows_by_group[g.group_name] = rows
    return rows_by_group


def _build_plan_type_rows(
    sections: list[_PlanSection], plan_charges: list[_PlanCharge],
) -> list[_DetailRow]:
    """One charged row per (plan, workload type), merged across groups, in plan-charge order.

    Only plans present in plan_charges are included, so a configured-only filter on
    the charges carries over to the breakdown.
    """
    buckets: dict[str, dict[str, _Tally]] = {}
    order: dict[tuple[str, str], int] = {}
    charge_ids = {c.plan_id for c in plan_charges}
    for s in sections:
        if s.plan_id not in charge_ids:
            continue
        per_type = buckets.setdefault(s.plan_id, {})
        for tr in s.rows:
            if tr.type_label not in per_type:
                per_type[tr.type_label] = _Tally()
                order[(s.plan_id, tr.type_label)] = tr.type_order
            per_type[tr.type_label].add(tr.count, tr.storage_bytes)
    rows: list[_DetailRow] = []
    for c in plan_charges:
        per_type = buckets.get(c.plan_id, {})
        pid = c.plan_id
        for label in sorted(per_type, key=lambda lbl: order[(pid, lbl)]):
            tally = per_type[label]
            storage_gb, instance_charge, storage_charge = _charge_fields(
                tally.count, tally.storage_bytes,
                c.charge_per_instance, c.charge_per_gb, gb_decimals=4,
            )
            rows.append(_DetailRow(
                plan_type=c.plan_type,
                plan_name=c.plan_name,
                plan_id=pid,
                pricing_plan=c.pricing_plan_name,
                workload_type=label,
                instances=tally.count,
                storage_gb=storage_gb,
                instance_charge=instance_charge,
                storage_charge=storage_charge,
            ))
    return rows


def _build_server_type_rows(
    server_stats: list[_ServerTypeStat], server_charges: list[_ServerCharge],
) -> list[_ServerDetailRow]:
    """One charged row per (server, workload type), in server-charge order.

    Only servers present in server_charges are included, so a configured-only filter
    on the charges carries over to the breakdown.
    """
    buckets: dict[str, dict[str, _Tally]] = {}
    order: dict[tuple[str, str], int] = {}
    charge_ns = {c.namespace for c in server_charges}
    for st in server_stats:
        if st.namespace not in charge_ns:
            continue
        per_type = buckets.setdefault(st.namespace, {})
        if st.type_label not in per_type:
            per_type[st.type_label] = _Tally()
            order[(st.namespace, st.type_label)] = st.type_order
        per_type[st.type_label].add(st.count, st.storage_bytes)
    rows: list[_ServerDetailRow] = []
    for c in server_charges:
        per_type = buckets.get(c.namespace, {})
        ns = c.namespace
        for label in sorted(per_type, key=lambda lbl: order[(ns, lbl)]):
            tally = per_type[label]
            storage_gb, instance_charge, storage_charge = _charge_fields(
                tally.count, tally.storage_bytes,
                c.charge_per_instance, c.charge_per_gb, gb_decimals=4,
            )
            rows.append(_ServerDetailRow(
                server_name=c.server_name,
                namespace=ns,
                pricing_plan=c.pricing_plan_name,
                workload_type=label,
                instances=tally.count,
                storage_gb=storage_gb,
                instance_charge=instance_charge,
                storage_charge=storage_charge,
            ))
    return rows


def _build_details_view(
    sections: list[_PlanSection],
    server_stats: list[_ServerTypeStat],
    plan_charges: list[_PlanCharge],
    server_charges: list[_ServerCharge],
    group_charges: list[_GroupCharge],
    server_names: dict[str, str],
) -> _DetailsView:
    """Assemble every --details row set from the granular scan data and charge rows."""
    return _DetailsView(
        server_rows_by_group=_build_group_server_rows(server_stats, group_charges, server_names),
        plan_rows_by_group=_build_group_plan_rows(sections, group_charges),
        workload_rows_by_group=_build_group_workload_rows(sections, group_charges),
        server_type_rows=_build_server_type_rows(server_stats, server_charges),
        plan_type_rows=_build_plan_type_rows(sections, plan_charges),
    )


# ── Output: table helpers ─────────────────────────────────────────────────────

def _kind_label(plan_type: str) -> str:
    if plan_type == _PROTECTION:
        return "Protection"
    if plan_type == _RETIREMENT:
        return "Retirement"
    return "Mixed"


def _billing_table_header(
    name_header: str, name_width: int, kind_header: str | None, pricing_width: int,
) -> str:
    kind_col = f"  {kind_header:<10}" if kind_header is not None else ""
    pricing_col = f"  {'Pricing Plan':<{pricing_width}}" if pricing_width else ""
    return (
        f"{name_header:<{name_width}}{kind_col}{pricing_col}"
        f"  {'Instances':>9}  {'Storage (GB)':>12}  {'Instance Chg':>13}  {'Storage Chg':>12}  {'Total Chg':>10}"
    )


def _billing_table_rule(name_width: int, has_kind: bool, pricing_width: int) -> str:
    kind_col = f"  {'-'*10}" if has_kind else ""
    pricing_col = f"  {'-'*pricing_width}" if pricing_width else ""
    return (
        f"{'-'*name_width}{kind_col}{pricing_col}"
        f"  {'-'*9}  {'-'*12}  {'-'*13}  {'-'*12}  {'-'*10}"
    )


def _billing_row(
    name: str, kind: str | None, pricing_plan_name: str,
    instances: int, storage_gb: float, instance_charge: float, storage_charge: float,
    name_width: int, pricing_width: int,
) -> str:
    total_charge = instance_charge + storage_charge
    kind_col = f"  {kind:<10}" if kind is not None else ""
    pricing_col = f"  {pricing_plan_name:<{pricing_width}}" if pricing_width else ""
    return (
        f"{name:<{name_width}}{kind_col}{pricing_col}"
        f"  {instances:>9}  {storage_gb:>12.2f}"
        f"  {fmt_money(instance_charge):>13}  {fmt_money(storage_charge):>12}  {fmt_money(total_charge):>10}"
    )


def _billing_total_row(
    charges: Sequence[_ChargedRow], has_kind: bool, name_width: int, pricing_width: int,
) -> None:
    """Print the closing rule + Total row of one billing table."""
    print(_billing_table_rule(name_width, has_kind, pricing_width))
    t = _aggregate(charges)
    print(_billing_row(
        "Total", "" if has_kind else None, "",
        t.instances, t.storage_gb, t.instance_charge, t.storage_charge,
        name_width, pricing_width,
    ))


# ── Output: table ─────────────────────────────────────────────────────────────

def _print_billing_table(
    charges: Sequence[_RatedCharge],
    rows: list[tuple[str, str | None]],
    name_header: str,
    kind_header: str | None,
    name_width: int,
    pricing_width: int,
) -> None:
    """Render one summary billing table; *rows* pairs each charge with its
    (name, kind label or None) display cells, in charge order."""
    print(_billing_table_header(name_header, name_width, kind_header, pricing_width))
    print(_billing_table_rule(name_width, kind_header is not None, pricing_width).replace("-", "═"))
    for c, (name, kind) in zip(charges, rows, strict=True):
        print(_billing_row(
            name, kind, c.pricing_plan_name,
            c.instances, c.storage_gb, c.instance_charge, c.storage_charge,
            name_width, pricing_width,
        ))
    _billing_total_row(charges, kind_header is not None, name_width, pricing_width)


def _print_pricing_plans_table(pricing: _PricingConfig) -> None:
    named_plans = [pp for pp in pricing.pricing_plans if pp.name]
    if not named_plans:
        return
    name_w = max(len("Name"), max(len(pp.name) for pp in named_plans))
    print("Pricing Plans")
    print(f"  {'Name':<{name_w}}  {'Rate/Instance':>13}  {'Rate/GB':>8}")
    print(f"  {'═'*name_w}  {'═'*13}  {'═'*8}")
    for pp in named_plans:
        print(f"  {pp.name:<{name_w}}  {fmt_money(pp.charge_per_instance):>13}  {fmt_money(pp.charge_per_gb):>8}")
    print()


# Text cells followed by numeric columns. Charged rows carry instances, storage_gb,
# instance_charge, storage_charge (total is derived); distribution rows carry only
# instances and storage_gb.
_DetailTableRow = tuple[list[str], int, float, float, float]
_DistTableRow = tuple[list[str], int, float]
_DetailRowT = TypeVar("_DetailRowT", _DetailRow, _ServerDetailRow, _GroupTypeRow)
_DistRowT = TypeVar("_DistRowT", _GroupServerRow, _GroupPlanRow)

_DETAIL_NUM_HEADER = (
    f"{'Instances':>9}  {'Storage (GB)':>12}  {'Instance Chg':>13}  {'Storage Chg':>12}  {'Total Chg':>10}"
)
_DETAIL_NUM_RULE = f"{'═'*9}  {'═'*12}  {'═'*13}  {'═'*12}  {'═'*10}"
_DIST_NUM_HEADER = f"{'Instances':>9}  {'Storage (GB)':>12}"
_DIST_NUM_RULE = f"{'═'*9}  {'═'*12}"


def _detail_table_rows(
    rows: list[_DetailRowT], cells: Callable[[_DetailRowT], list[str]],
) -> list[_DetailTableRow]:
    return [
        (cells(r), r.instances, r.storage_gb, r.instance_charge, r.storage_charge)
        for r in rows
    ]


def _dist_table_rows(
    rows: list[_DistRowT], cells: Callable[[_DistRowT], list[str]],
) -> list[_DistTableRow]:
    return [(cells(r), r.instances, r.storage_gb) for r in rows]


def _detail_col_widths(headers: list[str], cell_rows: list[list[str]]) -> list[int]:
    return [
        max(len(h), max((len(cells[i]) for cells in cell_rows), default=0))
        for i, h in enumerate(headers)
    ]


def _print_detail_table(
    headers: list[str],
    rows: list[_DetailTableRow],
    *,
    indent: str = "",
    widths: list[int],
) -> None:
    """Render one charged detail table: left-aligned text columns + five numeric columns.

    widths comes precomputed from _print_details_tables so all detail tables align.
    """
    text_header = "  ".join(f"{h:<{w}}" for h, w in zip(headers, widths, strict=True))
    print(f"{indent}{text_header}  {_DETAIL_NUM_HEADER}")
    text_rule = "  ".join("═" * w for w in widths)
    print(f"{indent}{text_rule}  {_DETAIL_NUM_RULE}")
    for cells, instances, storage_gb, instance_charge, storage_charge in rows:
        text = "  ".join(f"{c:<{w}}" for c, w in zip(cells, widths, strict=True))
        print(
            f"{indent}{text}  {instances:>9}  {storage_gb:>12.4f}"
            f"  {fmt_money(instance_charge):>13}  {fmt_money(storage_charge):>12}"
            f"  {fmt_money(instance_charge + storage_charge):>10}"
        )


def _print_dist_table(
    headers: list[str],
    rows: list[_DistTableRow],
    *,
    indent: str = "",
    widths: list[int],
) -> None:
    """Render one distribution table: left-aligned text columns + instances/storage only."""
    text_header = "  ".join(f"{h:<{w}}" for h, w in zip(headers, widths, strict=True))
    print(f"{indent}{text_header}  {_DIST_NUM_HEADER}")
    text_rule = "  ".join("═" * w for w in widths)
    print(f"{indent}{text_rule}  {_DIST_NUM_RULE}")
    for cells, instances, storage_gb in rows:
        text = "  ".join(f"{c:<{w}}" for c, w in zip(cells, widths, strict=True))
        print(f"{indent}{text}  {instances:>9}  {storage_gb:>12.2f}")


_GROUP_SERVER_HEADERS = ["Server"]
_GROUP_PLAN_HEADERS = ["Plan", "Plan Type"]
_GROUP_WORKLOAD_HEADERS = ["Workload Type"]
_SERVER_TYPE_HEADERS = ["Server", "Pricing Plan", "Workload Type"]
_PLAN_TYPE_HEADERS = ["Plan", "Plan Type", "Pricing Plan", "Workload Type"]


def _detail_text_width(widths: list[int]) -> int:
    """Width of a table's text-column section, including per-column trailing separators."""
    return sum(widths) + 2 * len(widths)


def _print_group_detail_blocks(
    group_charges: list[_GroupCharge],
    server_rows_by_group: dict[str, list[_DistTableRow]],
    plan_rows_by_group: dict[str, list[_DistTableRow]],
    workload_rows_by_group: dict[str, list[_DetailTableRow]],
    server_widths: list[int],
    plan_widths: list[int],
    workload_widths: list[int],
) -> None:
    """One block per group: server and plan distributions, then charged workload rows."""
    for g in group_charges:
        title = f"Group: {g.group_name}"
        if g.pricing_plan_name:
            title += f"  (Pricing Plan: {g.pricing_plan_name})"
        print(title)
        print()
        print("  Backup Servers")
        _print_dist_table(
            _GROUP_SERVER_HEADERS, server_rows_by_group[g.group_name],
            indent="  ", widths=server_widths,
        )
        print()
        print("  Plans")
        _print_dist_table(
            _GROUP_PLAN_HEADERS, plan_rows_by_group[g.group_name],
            indent="  ", widths=plan_widths,
        )
        print()
        print("  Workloads")
        _print_detail_table(
            _GROUP_WORKLOAD_HEADERS, workload_rows_by_group[g.group_name],
            indent="  ", widths=workload_widths,
        )
        print()
        print(
            f"  Group Total: {g.instances} instances, "
            f"{g.storage_gb:,.2f} GB, {fmt_money(g.total_charge)}"
        )
        print()


def _print_details_tables(
    group_charges: list[_GroupCharge],
    details: _DetailsView,
    pricing: _PricingConfig,
    only: str,
) -> None:
    _print_pricing_plans_table(pricing)

    show_groups = only in ("", "groups") and bool(group_charges)
    show_servers = only in ("", "servers")
    show_plans = only in ("", "plans")

    group_server_rows = {
        name: _dist_table_rows(rows, lambda r: [r.server_name])
        for name, rows in details.server_rows_by_group.items()
    } if show_groups else {}
    group_plan_rows = {
        name: _dist_table_rows(rows, lambda r: [r.plan_name, _kind_label(r.plan_type)])
        for name, rows in details.plan_rows_by_group.items()
    } if show_groups else {}
    group_workload_rows = {
        name: _detail_table_rows(rows, lambda r: [r.workload_type])
        for name, rows in details.workload_rows_by_group.items()
    } if show_groups else {}
    server_type_rows = _detail_table_rows(
        details.server_type_rows,
        lambda r: [r.server_name, r.pricing_plan, r.workload_type],
    ) if show_servers else []
    plan_type_rows = _detail_table_rows(
        details.plan_type_rows,
        lambda r: [r.plan_name, _kind_label(r.plan_type), r.pricing_plan, r.workload_type],
    ) if show_plans else []

    # Equal total width for every detail table: the indent is identical everywhere but
    # the numeric section differs (charged vs distribution), so pad each table's first
    # column until text section + numeric section matches the widest table that will
    # actually print. Group tables use rows merged across groups so all group blocks
    # share one width set.
    charged_num = 2 + len(_DETAIL_NUM_HEADER)
    dist_num = 2 + len(_DIST_NUM_HEADER)
    tables: list[tuple[list[str], list[list[str]], int]] = [
        (_GROUP_SERVER_HEADERS, [r[0] for rows in group_server_rows.values() for r in rows], dist_num),
        (_GROUP_PLAN_HEADERS, [r[0] for rows in group_plan_rows.values() for r in rows], dist_num),
        (_GROUP_WORKLOAD_HEADERS,
         [r[0] for rows in group_workload_rows.values() for r in rows], charged_num),
        (_SERVER_TYPE_HEADERS, [r[0] for r in server_type_rows], charged_num),
        (_PLAN_TYPE_HEADERS, [r[0] for r in plan_type_rows], charged_num),
    ]
    width_sets: list[list[int] | None] = [
        _detail_col_widths(headers, cell_rows) if cell_rows else None
        for headers, cell_rows, _ in tables
    ]
    target = max(
        (_detail_text_width(ws) + num for (_, _, num), ws in zip(tables, width_sets, strict=True) if ws is not None),
        default=0,
    )
    for (_, _, num), ws in zip(tables, width_sets, strict=True):
        if ws is not None:
            ws[0] += target - (_detail_text_width(ws) + num)
    gs_widths, gp_widths, gw_widths, st_widths, pt_widths = width_sets

    if show_groups:
        # Groups always carry server, plan, and workload rows, so their widths exist.
        assert gs_widths is not None and gp_widths is not None and gw_widths is not None
        _print_group_detail_blocks(
            group_charges, group_server_rows, group_plan_rows, group_workload_rows,
            gs_widths, gp_widths, gw_widths,
        )
    if server_type_rows:
        assert st_widths is not None
        print("Backup Servers by Workload Type")
        _print_detail_table(_SERVER_TYPE_HEADERS, server_type_rows, indent="  ", widths=st_widths)
        print()
    if plan_type_rows:
        assert pt_widths is not None
        print("Plans by Workload Type")
        _print_detail_table(_PLAN_TYPE_HEADERS, plan_type_rows, indent="  ", widths=pt_widths)
        print()


def _print_table(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    pricing: _PricingConfig,
    details: _DetailsView | None,
    only: str,
) -> None:
    show_groups = only in ("", "groups") and bool(group_charges)
    show_servers = only in ("", "servers") and bool(server_charges)
    show_plans = only in ("", "plans") and bool(plan_charges)
    if not (show_groups or show_servers or show_plans):
        print("(no workloads)")
        return

    all_names = (
        ([g.pricing_plan_name for g in group_charges] if show_groups else [])
        + ([c.pricing_plan_name for c in server_charges] if show_servers else [])
        + ([c.pricing_plan_name for c in plan_charges] if show_plans else [])
    )
    has_pricing = any(all_names)
    pricing_width = max(len("Pricing Plan"), max(len(n) for n in all_names)) if has_pricing else 0
    extra = (2 + pricing_width) if pricing_width else 0

    # Compute a unified left-column width so all rendered tables have equal total width.
    group_min  = max(len("Group"),  max(len(g.group_name)  for g in group_charges))  if show_groups  else 0
    server_min = max(len("Server"), max(len(c.server_name) for c in server_charges)) if show_servers else 0
    plan_min   = max(len("Plan"),   max(len(c.plan_name)   for c in plan_charges))   if show_plans   else 0
    L = max(group_min + extra, server_min + extra, plan_min + 12 + extra)

    if show_groups:
        _print_billing_table(
            group_charges, [(g.group_name, None) for g in group_charges],
            "Group", None, L - extra, pricing_width,
        )
        print()
    if show_servers:
        _print_billing_table(
            server_charges, [(c.server_name, None) for c in server_charges],
            "Server", None, L - extra, pricing_width,
        )
        print()
    if show_plans:
        _print_billing_table(
            plan_charges, [(c.plan_name, _kind_label(c.plan_type)) for c in plan_charges],
            "Plan", "Plan Type", L - 12 - extra, pricing_width,
        )
        print()

    if details is not None:
        _print_details_tables(group_charges, details, pricing, only)


# ── Output: shared detail columns (CSV + XLSX) ───────────────────────────────

# Column names double as attribute names on the detail-row dataclasses (see
# _detail_values). Shared columns keep the same relative order as the table headers.
_GROUP_SERVER_COLS = ["group_name", "server_name", "namespace", "instances", "storage_gb"]
_GROUP_PLAN_COLS = ["group_name", "plan_name", "plan_id", "plan_type", "instances", "storage_gb"]
_GROUP_WORKLOAD_COLS = [
    "group_name", "pricing_plan", "workload_type", "instances", "storage_gb",
    "instance_charge", "storage_charge", "total_charge",
]
_SERVER_TYPE_COLS = [
    "server_name", "namespace", "pricing_plan",
    "workload_type", "instances", "storage_gb",
    "instance_charge", "storage_charge", "total_charge",
]
_PLAN_TYPE_COLS = [
    "plan_name", "plan_id", "plan_type", "pricing_plan",
    "workload_type", "instances", "storage_gb",
    "instance_charge", "storage_charge", "total_charge",
]

_DetailValueRow = _DetailRow | _ServerDetailRow | _GroupTypeRow | _GroupServerRow | _GroupPlanRow


def _detail_values(r: _DetailValueRow, cols: list[str]) -> list[Any]:
    """Row values for cols, in order; total_charge (a derived property) is rounded."""
    return [round(getattr(r, c), 2) if c == "total_charge" else getattr(r, c) for c in cols]


class _DetailSection(NamedTuple):
    # title doubles as the XLSX sheet name — Excel caps sheet names at 31 characters
    # ("Backup Servers by Workload Type" is exactly 31).
    title: str
    cols: list[str]
    values: list[list[Any]]


def _detail_sections(details: _DetailsView, only: str) -> list[_DetailSection]:
    """The detail sections in display order: the three group breakdowns, then the
    per-workload-type breakdowns of the Backup Servers and Plans dimensions. Grouped
    rows iterate the by-group dicts so row order matches the table blocks.
    """
    sections: list[_DetailSection] = []
    if only in ("", "groups"):
        grouped_servers = [r for rows in details.server_rows_by_group.values() for r in rows]
        grouped_plans = [r for rows in details.plan_rows_by_group.values() for r in rows]
        grouped_workloads = [r for rows in details.workload_rows_by_group.values() for r in rows]
        sections += [
            _DetailSection("Group Backup Servers", _GROUP_SERVER_COLS,
                           [_detail_values(r, _GROUP_SERVER_COLS) for r in grouped_servers]),
            _DetailSection("Group Plans", _GROUP_PLAN_COLS,
                           [_detail_values(r, _GROUP_PLAN_COLS) for r in grouped_plans]),
            _DetailSection("Group Workloads", _GROUP_WORKLOAD_COLS,
                           [_detail_values(r, _GROUP_WORKLOAD_COLS) for r in grouped_workloads]),
        ]
    if only in ("", "servers"):
        sections.append(_DetailSection(
            "Backup Servers by Workload Type", _SERVER_TYPE_COLS,
            [_detail_values(r, _SERVER_TYPE_COLS) for r in details.server_type_rows]))
    if only in ("", "plans"):
        sections.append(_DetailSection(
            "Plans by Workload Type", _PLAN_TYPE_COLS,
            [_detail_values(r, _PLAN_TYPE_COLS) for r in details.plan_type_rows]))
    return sections


class _SummarySection(NamedTuple):
    """One summary billing table (Groups / Backup Servers / Plans) for CSV and XLSX."""
    title: str  # CSV section title and XLSX sheet name
    cols: list[str]
    rows: list[list[Any]]
    total: list[Any]  # "Total" + blank pads + the five aggregate values


def _summary_sections(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    only: str,
) -> list[_SummarySection]:
    """The non-empty summary sections selected by *only*, in display order."""

    def _total(cols: list[str], charges: Sequence[_RatedCharge]) -> list[Any]:
        t = _aggregate(charges)
        return ["Total"] + [""] * (len(cols) - 6) + [
            t.instances, t.storage_gb, t.instance_charge, t.storage_charge, t.total_charge,
        ]

    sections: list[_SummarySection] = []
    if only in ("", "groups") and group_charges:
        cols = ["group", "kind", "pricing_plan", "charge_per_instance", "charge_per_gb",
                "instances", "storage_gb", "instance_charge", "storage_charge", "total_charge"]
        rows = [
            [g.group_name, g.plan_type, g.pricing_plan_name,
             g.charge_per_instance, g.charge_per_gb,
             g.instances, g.storage_gb, g.instance_charge, g.storage_charge,
             round(g.total_charge, 2)]
            for g in group_charges
        ]
        sections.append(_SummarySection("Groups", cols, rows, _total(cols, group_charges)))
    if only in ("", "servers") and server_charges:
        cols = ["server_name", "namespace", "pricing_plan", "charge_per_instance", "charge_per_gb",
                "instances", "storage_gb", "instance_charge", "storage_charge", "total_charge"]
        rows = [
            [c.server_name, c.namespace, c.pricing_plan_name,
             c.charge_per_instance, c.charge_per_gb,
             c.instances, c.storage_gb, c.instance_charge, c.storage_charge,
             round(c.total_charge, 2)]
            for c in server_charges
        ]
        sections.append(_SummarySection("Backup Servers", cols, rows, _total(cols, server_charges)))
    if only in ("", "plans") and plan_charges:
        cols = ["plan", "plan_id", "kind", "pricing_plan", "charge_per_instance", "charge_per_gb",
                "instances", "storage_gb", "instance_charge", "storage_charge", "total_charge"]
        rows = [
            [pc.plan_name, pc.plan_id, pc.plan_type, pc.pricing_plan_name,
             pc.charge_per_instance, pc.charge_per_gb,
             pc.instances, pc.storage_gb, pc.instance_charge, pc.storage_charge,
             round(pc.total_charge, 2)]
            for pc in plan_charges
        ]
        sections.append(_SummarySection("Plans", cols, rows, _total(cols, plan_charges)))
    return sections


# ── Output: CSV ──────────────────────────────────────────────────────────────

def _print_csv(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    pricing: _PricingConfig,
    details: _DetailsView | None,
    only: str,
) -> None:
    w = csv.writer(sys.stdout)
    first_section = True

    def _section_title(title: str) -> None:
        nonlocal first_section
        if not first_section:
            w.writerow([])
        first_section = False
        w.writerow([title])

    for summary in _summary_sections(plan_charges, group_charges, server_charges, only):
        _section_title(summary.title)
        w.writerow(summary.cols)
        w.writerows(summary.rows)
        w.writerow(summary.total)

    if details is None:
        return

    _section_title("Pricing Plans")
    w.writerow(["name", "charge_per_instance", "charge_per_gb"])
    for pp in pricing.pricing_plans:
        w.writerow([pp.name, pp.charge_per_instance, pp.charge_per_gb])

    for section in _detail_sections(details, only):
        if not section.values:
            continue
        _section_title(section.title)
        w.writerow(section.cols)
        w.writerows(section.values)


# ── Output: JSON ─────────────────────────────────────────────────────────────

def _print_json(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    pricing: _PricingConfig,
    details: _DetailsView | None,
    only: str,
) -> None:
    show_groups = only in ("", "groups")
    show_servers = only in ("", "servers")
    show_plans = only in ("", "plans")

    def _pricing_obj(name: str, per_instance: float, per_gb: float) -> dict[str, Any]:
        return {"name": name, "charge_per_instance": per_instance, "charge_per_gb": per_gb}

    def _charge_obj(c: _RatedCharge) -> dict[str, Any]:
        """The rate-card and charge keys shared by group/server/plan entries."""
        return {
            "pricing_plan": _pricing_obj(
                c.pricing_plan_name, c.charge_per_instance, c.charge_per_gb
            ),
            "instances": c.instances,
            "storage_gb": c.storage_gb,
            "instance_charge": c.instance_charge,
            "storage_charge": c.storage_charge,
            "total_charge": round(c.total_charge, 2),
        }

    def _type_obj(r: _DetailRow | _ServerDetailRow | _GroupTypeRow) -> dict[str, Any]:
        return {
            "workload_type": r.workload_type,
            "instances": r.instances,
            "storage_gb": r.storage_gb,
            "instance_charge": r.instance_charge,
            "storage_charge": r.storage_charge,
            "total_charge": round(r.total_charge, 2),
        }

    out: dict[str, Any] = {
        "pricing_plans": [
            _pricing_obj(pp.name, pp.charge_per_instance, pp.charge_per_gb)
            for pp in pricing.pricing_plans
        ],
        "groups": [
            {"group_name": g.group_name, "plan_type": g.plan_type, **_charge_obj(g)}
            for g in group_charges
        ] if show_groups else [],
        "backup_servers": [
            {"server_name": c.server_name, "namespace": c.namespace, **_charge_obj(c)}
            for c in server_charges
        ] if show_servers else [],
        "plans": [
            {"plan_name": c.plan_name, "plan_type": c.plan_type, "plan_id": c.plan_id,
             **_charge_obj(c)}
            for c in plan_charges
        ] if show_plans else [],
    }
    totals: dict[str, Any] = {}
    if show_groups:
        totals["groups"] = asdict(_aggregate(group_charges))
    if show_servers:
        totals["backup_servers"] = asdict(_aggregate(server_charges))
    if show_plans:
        totals["plans"] = asdict(_aggregate(plan_charges))
    out["totals"] = totals

    if details is not None:
        if show_groups:
            for group_dict, g in zip(out["groups"], group_charges, strict=True):
                group_dict["backup_servers"] = [
                    {
                        "server_name": r.server_name,
                        "namespace": r.namespace,
                        "instances": r.instances,
                        "storage_gb": r.storage_gb,
                    }
                    for r in details.server_rows_by_group.get(g.group_name, [])
                ]
                group_dict["plans"] = [
                    {
                        "plan_name": r.plan_name,
                        "plan_type": r.plan_type,
                        "plan_id": r.plan_id,
                        "instances": r.instances,
                        "storage_gb": r.storage_gb,
                    }
                    for r in details.plan_rows_by_group.get(g.group_name, [])
                ]
                group_dict["workloads"] = [
                    _type_obj(r) for r in details.workload_rows_by_group.get(g.group_name, [])
                ]
        if show_servers:
            rows_by_ns: dict[str, list[_ServerDetailRow]] = {}
            for sr in details.server_type_rows:
                rows_by_ns.setdefault(sr.namespace, []).append(sr)
            for server_dict, sc in zip(out["backup_servers"], server_charges, strict=True):
                server_dict["by_type"] = [_type_obj(r) for r in rows_by_ns.get(sc.namespace, [])]
        if show_plans:
            rows_by_plan: dict[str, list[_DetailRow]] = {}
            for pr in details.plan_type_rows:
                rows_by_plan.setdefault(pr.plan_id, []).append(pr)
            for plan_dict, pc in zip(out["plans"], plan_charges, strict=True):
                plan_dict["by_type"] = [_type_obj(r) for r in rows_by_plan.get(pc.plan_id, [])]

    print(json.dumps(out, indent=2))


# ── Output: XLSX ──────────────────────────────────────────────────────────────

_FMT_CURRENCY = "#,##0.00"
_FMT_GB2 = "#,##0.00"
_FMT_GB4 = "#,##0.0000"
_BOLD = openpyxl.styles.Font(bold=True)


def _xlsx_header(ws: _Worksheet, cols: list[str]) -> None:
    ws.append(cols)
    for cell in ws[1]:
        cell.font = _BOLD


def _xlsx_total_row(ws: _Worksheet, values: list[Any]) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def _xlsx_fmt(
    ws: _Worksheet,
    row_idx: int,
    col_indices: list[int],
    fmt: str,
) -> None:
    for col in col_indices:
        ws.cell(row=row_idx, column=col).number_format = fmt


def _xlsx_autofit(ws: _Worksheet) -> None:
    for idx, col_cells in enumerate(ws.columns):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[_get_col_letter(idx + 1)].width = max_len + 2


def _xlsx_summary_sheet(wb: openpyxl.Workbook, section: _SummarySection) -> None:
    """One summary sheet (rows + bold Total row); number-format columns are
    located by name in section.cols."""
    ws = wb.create_sheet(section.title)
    _xlsx_header(ws, section.cols)
    for row_values in section.rows:
        ws.append(row_values)
    _xlsx_total_row(ws, section.total)
    money_cols = [
        section.cols.index(c) + 1
        for c in ("charge_per_instance", "charge_per_gb",
                  "instance_charge", "storage_charge", "total_charge")
    ]
    gb_cols = [section.cols.index("storage_gb") + 1]
    for row_idx in range(2, ws.max_row + 1):
        _xlsx_fmt(ws, row_idx, money_cols, _FMT_CURRENCY)
        _xlsx_fmt(ws, row_idx, gb_cols, _FMT_GB2)
    _xlsx_autofit(ws)


def _xlsx_detail_sheet(
    wb: openpyxl.Workbook, title: str, cols: list[str], values: list[list[Any]],
) -> None:
    """One detail sheet; number-format columns are located by name in cols.

    Distribution sections carry no charge columns and 2-decimal storage aggregates;
    charged sections use the 4-decimal per-type storage format.
    """
    ws = wb.create_sheet(title)
    _xlsx_header(ws, cols)
    for row_values in values:
        ws.append(row_values)
    money_cols = [
        cols.index(c) + 1
        for c in ("instance_charge", "storage_charge", "total_charge")
        if c in cols
    ]
    gb_cols = [cols.index("storage_gb") + 1]
    gb_fmt = _FMT_GB4 if money_cols else _FMT_GB2
    for row_idx in range(2, ws.max_row + 1):
        _xlsx_fmt(ws, row_idx, money_cols, _FMT_CURRENCY)
        _xlsx_fmt(ws, row_idx, gb_cols, gb_fmt)
    _xlsx_autofit(ws)


def _write_xlsx(
    plan_charges: list[_PlanCharge],
    group_charges: list[_GroupCharge],
    server_charges: list[_ServerCharge],
    pricing: _PricingConfig,
    details: _DetailsView | None,
    only: str,
    output_file: str,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    for summary in _summary_sections(plan_charges, group_charges, server_charges, only):
        _xlsx_summary_sheet(wb, summary)

    if details is not None:
        # ── Pricing Plans ─────────────────────────────────────────────────────
        named_plans = [pp for pp in pricing.pricing_plans if pp.name]
        if named_plans:
            ws = wb.create_sheet("Pricing Plans")
            _xlsx_header(ws, ["name", "charge_per_instance", "charge_per_gb"])
            for pp in named_plans:
                ws.append([pp.name, pp.charge_per_instance, pp.charge_per_gb])
            for row_idx in range(2, ws.max_row + 1):
                _xlsx_fmt(ws, row_idx, [2, 3], _FMT_CURRENCY)
            _xlsx_autofit(ws)

        # ── Detail sheets ─────────────────────────────────────────────────────
        for section in _detail_sections(details, only):
            if section.values:
                _xlsx_detail_sheet(wb, section.title, section.cols, section.values)

    if not wb.sheetnames:  # a workbook must contain at least one sheet
        wb.create_sheet("Report")

    wb.save(output_file)
    print(f"Saved: {output_file}", file=sys.stderr)


# ── Entry point ──────────────────────────────────────────────────────────────

async def run(
    *,
    output_format: str,
    pricing: _PricingConfig,
    show_details: bool,
    concurrency: int,
    output_file: str | None,
    only: str,
    configured_only: bool,
    profile: str | None = None,
) -> None:
    print("Collecting data...", file=sys.stderr)
    async with make_client(profile=profile) as apm:
        servers = await collect_backup_servers(apm)
        ns_by_server_id = {s.backup_server_id: s.namespace for s in servers}
        server_names = {s.namespace: s.name for s in servers}
        unknown_ids = pricing.resolve_server_ids(ns_by_server_id)
        if unknown_ids:
            print(
                "Warning: pricing config references unknown backup server ID(s): "
                + ", ".join(unknown_ids),
                file=sys.stderr,
            )
        sections, server_stats = await _scan_billing(apm, concurrency=concurrency, pricing=pricing)

    plan_charges = _compute_plan_charges(sections, pricing)
    group_charges = _compute_group_charges(sections, pricing)
    server_charges = _compute_server_charges(server_stats, pricing, server_names)

    if only == "groups" and sections and not group_charges:
        print(
            "Warning: no workloads matched any group; "
            "check the plans/backup_servers in the pricing config.",
            file=sys.stderr,
        )
    if configured_only:
        plan_charges = [c for c in plan_charges if c.plan_id in pricing.configured_plan_ids]
        server_charges = [c for c in server_charges if c.namespace in pricing.configured_namespaces]

    details = None
    if show_details:
        details = _build_details_view(
            sections, server_stats, plan_charges, server_charges, group_charges, server_names,
        )

    if output_format == "csv":
        _print_csv(plan_charges, group_charges, server_charges, pricing, details, only)
    elif output_format == "json":
        _print_json(plan_charges, group_charges, server_charges, pricing, details, only)
    elif output_format == "xlsx":
        assert output_file is not None
        _write_xlsx(plan_charges, group_charges, server_charges, pricing, details, only, output_file)
    else:
        _print_table(plan_charges, group_charges, server_charges, pricing, details, only)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--config", metavar="FILE",
        help=(
            "Path to a YAML pricing config file. Defines named pricing plans, groups "
            "(membership by APM plan and/or backup server ID; groups may overlap), and "
            "per-server / per-plan "
            "pricing assignments. When given, the Backup Servers and Plans sections list only "
            "configured items unless --show-not-configured is passed. Generate a template with "
            "--dump-config-template. Cannot be combined with --charge-per-instance or "
            "--charge-per-gb."
        ),
    )
    parser.add_argument(
        "--dump-config-template", dest="dump_config_template", action="store_true",
        help="Print a commented YAML pricing config template to stdout and exit.",
    )
    parser.add_argument(
        "--charge-per-instance", dest="charge_per_instance", type=float, default=0.0,
        help=(
            "Monetary charge per workload instance, applied uniformly to all servers and plans "
            "(default: 0). Cannot be combined with --config."
        ),
    )
    parser.add_argument(
        "--charge-per-gb", dest="charge_per_gb", type=float, default=0.0,
        help=(
            "Monetary charge per GB of storage usage, applied uniformly to all servers and plans "
            "(default: 0). Cannot be combined with --config."
        ),
    )
    parser.add_argument(
        "--only", choices=["groups", "servers", "plans"], default="",
        help=(
            "Print a single billing section: groups, servers (backup servers), or plans "
            "(default: all three). The sections are independent views "
            "over the same workloads, so this only selects what is shown."
        ),
    )
    parser.add_argument(
        "--show-not-configured", dest="show_not_configured", action="store_true",
        help=(
            "Also list servers/plans not mentioned in the config (they use the fallback "
            "pricing plan). By default with --config, the Backup Servers and Plans sections "
            "list only servers/plans mentioned in the config (as a rate assignment or group "
            "member); the Groups section is unaffected. Requires --config."
        ),
    )
    parser.add_argument(
        "--details", dest="details", action="store_true",
        help=(
            "Print a per-workload-type breakdown: one block per group (its backup servers "
            "and plans as charge-free distributions, plus charged per-workload-type rows), "
            "followed by per-workload-type sections for the Backup Servers and Plans "
            "dimensions (default: totals only)."
        ),
    )
    parser.add_argument(
        "--concurrency", type=int, default=_DEFAULT_CONCURRENCY, metavar="N",
        help=f"Max concurrent API calls in flight while fetching workloads (default: {_DEFAULT_CONCURRENCY}).",
    )
    parser.add_argument(
        "-o", "--output", dest="output", choices=["table", "csv", "json", "xlsx"], default="table",
        help="Output format: table, csv, json, or xlsx (default: table).",
    )
    parser.add_argument(
        "--output-file", dest="output_file", metavar="FILE",
        help="Path for the output .xlsx file. Required when -o xlsx.",
    )
    add_profile_arg(parser)
    args = parser.parse_args()

    if args.output == "xlsx" and not args.output_file:
        parser.error("-o xlsx requires --output-file FILE")

    if args.dump_config_template:
        _dump_config_template()
        sys.exit(0)

    if args.config and (args.charge_per_instance or args.charge_per_gb):
        parser.error("--config cannot be combined with --charge-per-instance or --charge-per-gb")

    if args.show_not_configured and not args.config:
        parser.error("--show-not-configured requires --config")

    if args.config:
        try:
            pricing = _load_pricing_yaml(args.config)
        except (OSError, ValueError, KeyError) as e:
            print(f"Error loading pricing file: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        pricing = _PricingConfig(pricing_plans=[
            _PricingPlan(
                name="",
                charge_per_instance=args.charge_per_instance,
                charge_per_gb=args.charge_per_gb,
            ),
        ])

    if args.only == "groups" and not pricing.groups:
        parser.error("--only groups requires a --config file that defines groups")

    run_main(run(
        output_format=args.output,
        pricing=pricing,
        show_details=args.details,
        concurrency=args.concurrency,
        output_file=args.output_file,
        only=args.only,
        configured_only=bool(args.config) and not args.show_not_configured,
        profile=args.profile,
    ))


if __name__ == "__main__":
    main()
